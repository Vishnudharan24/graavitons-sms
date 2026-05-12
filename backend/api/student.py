from fastapi import FastAPI, HTTPException, status, UploadFile, File, Form, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, validator
from typing import List, Optional
import psycopg2
from psycopg2 import sql
from datetime import datetime, date
import pandas as pd
import io
from config import CORS_ORIGINS, APP_TITLE
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from api.middleware import get_current_user
from db_pool import get_db_connection
import os
import shutil

app = FastAPI(title=APP_TITLE)


def safe_float(value):
    if value is None or str(value).strip() == '':
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None

def safe_int(value):
    """Safely convert any value (including numpy floats) to Python int"""
    if value is None:
        return None
    try:
        return int(float(str(value)))
    except (ValueError, TypeError):
        return None


def safe_str(value):
    """Safely convert any value to a clean Python string"""
    if value is None:
        return None
    s = str(value)
    # Remove trailing .0 from numeric strings (e.g. "9876543210.0" -> "9876543210")
    if s.endswith('.0') and s[:-2].replace('-', '', 1).isdigit():
        s = s[:-2]
    return s

# CORS configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# Pydantic models
class StudentCreate(BaseModel):
    # Student table fields
    student_id: str
    batch_id: int
    student_name: str
    dob: Optional[date] = None
    grade: Optional[str] = None
    community: Optional[str] = None
    enrollment_year: Optional[int] = None
    course: Optional[str] = None
    board: Optional[str] = None
    gender: Optional[str] = None
    student_mobile: Optional[str] = None
    aadhar_no: Optional[str] = None
    apaar_id: Optional[str] = None
    email: Optional[str] = None
    photo_url: Optional[str] = None
    school_name: Optional[str] = None
    
    # Parent info fields
    guardian_name: Optional[str] = None
    guardian_occupation: Optional[str] = None
    guardian_mobile: Optional[str] = None
    guardian_email: Optional[str] = None
    father_name: Optional[str] = None
    father_occupation: Optional[str] = None
    father_mobile: Optional[str] = None
    father_email: Optional[str] = None
    mother_name: Optional[str] = None
    mother_occupation: Optional[str] = None
    mother_mobile: Optional[str] = None
    mother_email: Optional[str] = None
    sibling_name: Optional[str] = None
    sibling_grade: Optional[str] = None
    sibling_school: Optional[str] = None
    sibling_college: Optional[str] = None
    
    # 10th marks
    tenth_school_name: Optional[str] = None
    tenth_year_of_passing: Optional[int] = None
    tenth_board_of_study: Optional[str] = None
    tenth_english: Optional[int] = None
    tenth_tamil: Optional[int] = None
    tenth_hindi: Optional[int] = None
    tenth_maths: Optional[int] = None
    tenth_science: Optional[int] = None
    tenth_social_science: Optional[int] = None
    tenth_total_marks: Optional[int] = None
    
    # entrance exams
    "entrance_exam_1", "entrance_exam_1_percentile", "entrance_exam_1_mark",
    "entrance_exam_2", "entrance_exam_2_percentile", "entrance_exam_2_mark",
    "entrance_exam_3", "entrance_exam_3_percentile", "entrance_exam_3_mark",
    # 12th marks
    twelfth_school_name: Optional[str] = None
    twelfth_year_of_passing: Optional[int] = None
    twelfth_board_of_study: Optional[str] = None
    twelfth_english: Optional[int] = None
    twelfth_tamil: Optional[int] = None
    twelfth_physics: Optional[int] = None
    twelfth_chemistry: Optional[int] = None
    twelfth_maths: Optional[int] = None
    twelfth_biology: Optional[int] = None
    twelfth_computer_science: Optional[int] = None
    twelfth_total_marks: Optional[int] = None
    
    # Entrance exams
    entrance_exam_1: Optional[str] = None
    entrance_exam_1_percentile: Optional[float] = None
    entrance_exam_1_mark: Optional[int] = None
    entrance_exam_2: Optional[str] = None
    entrance_exam_2_percentile: Optional[float] = None
    entrance_exam_2_mark: Optional[int] = None
    entrance_exam_3: Optional[str] = None
    entrance_exam_3_percentile: Optional[float] = None
    entrance_exam_3_mark: Optional[int] = None
    
    # Entrance exams
    entrance_exam_1: Optional[str] = None
    entrance_exam_1_percentile: Optional[float] = None
    entrance_exam_1_mark: Optional[int] = None
    entrance_exam_2: Optional[str] = None
    entrance_exam_2_percentile: Optional[float] = None
    entrance_exam_2_mark: Optional[int] = None
    entrance_exam_3: Optional[str] = None
    entrance_exam_3_percentile: Optional[float] = None
    entrance_exam_3_mark: Optional[int] = None
    
    # Counselling details
    counselling_forum_1: Optional[str] = None
    counselling_round_1: Optional[int] = None
    all_india_rank_1: Optional[int] = None
    community_rank_1: Optional[int] = None
    counselling_college_1: Optional[str] = None
    counselling_forum_2: Optional[str] = None
    counselling_round_2: Optional[int] = None
    all_india_rank_2: Optional[int] = None
    community_rank_2: Optional[int] = None
    counselling_college_2: Optional[str] = None
    counselling_forum_3: Optional[str] = None
    counselling_round_3: Optional[int] = None
    all_india_rank_3: Optional[int] = None
    community_rank_3: Optional[int] = None
    counselling_college_3: Optional[str] = None


class StudentResponse(BaseModel):
    student_id: str
    student_name: str
    batch_id: int
    grade: Optional[str]
    gender: Optional[str]
    email: Optional[str]
    created_at: Optional[datetime]


class StudentUpdate(BaseModel):
    """Model for updating student - all fields optional for partial updates"""
    # Student table fields
    student_name: Optional[str] = None
    dob: Optional[date] = None
    grade: Optional[str] = None
    community: Optional[str] = None
    enrollment_year: Optional[int] = None
    course: Optional[str] = None
    board: Optional[str] = None
    gender: Optional[str] = None
    student_mobile: Optional[str] = None
    aadhar_no: Optional[str] = None
    apaar_id: Optional[str] = None
    email: Optional[str] = None
    school_name: Optional[str] = None
    
    # Parent info
    guardian_name: Optional[str] = None
    guardian_occupation: Optional[str] = None
    guardian_mobile: Optional[str] = None
    guardian_email: Optional[str] = None
    father_name: Optional[str] = None
    father_occupation: Optional[str] = None
    father_mobile: Optional[str] = None
    father_email: Optional[str] = None
    mother_name: Optional[str] = None
    mother_occupation: Optional[str] = None
    mother_mobile: Optional[str] = None
    mother_email: Optional[str] = None
    sibling_name: Optional[str] = None
    sibling_grade: Optional[str] = None
    sibling_school: Optional[str] = None
    sibling_college: Optional[str] = None
    
    # 10th marks
    tenth_school_name: Optional[str] = None
    tenth_year_of_passing: Optional[int] = None
    tenth_board_of_study: Optional[str] = None
    tenth_english: Optional[int] = None
    tenth_tamil: Optional[int] = None
    tenth_hindi: Optional[int] = None
    tenth_maths: Optional[int] = None
    tenth_science: Optional[int] = None
    tenth_social_science: Optional[int] = None
    tenth_total_marks: Optional[int] = None
    
    # entrance exams
    "entrance_exam_1", "entrance_exam_1_percentile", "entrance_exam_1_mark",
    "entrance_exam_2", "entrance_exam_2_percentile", "entrance_exam_2_mark",
    "entrance_exam_3", "entrance_exam_3_percentile", "entrance_exam_3_mark",
    # 12th marks
    twelfth_school_name: Optional[str] = None
    twelfth_year_of_passing: Optional[int] = None
    twelfth_board_of_study: Optional[str] = None
    twelfth_english: Optional[int] = None
    twelfth_tamil: Optional[int] = None
    twelfth_physics: Optional[int] = None
    twelfth_chemistry: Optional[int] = None
    twelfth_maths: Optional[int] = None
    twelfth_biology: Optional[int] = None
    twelfth_computer_science: Optional[int] = None
    twelfth_total_marks: Optional[int] = None
    
    # Counselling details
    counselling_forum_1: Optional[str] = None
    counselling_round_1: Optional[int] = None
    all_india_rank_1: Optional[int] = None
    community_rank_1: Optional[int] = None
    counselling_college_1: Optional[str] = None
    counselling_forum_2: Optional[str] = None
    counselling_round_2: Optional[int] = None
    all_india_rank_2: Optional[int] = None
    community_rank_2: Optional[int] = None
    counselling_college_2: Optional[str] = None
    counselling_forum_3: Optional[str] = None
    counselling_round_3: Optional[int] = None
    all_india_rank_3: Optional[int] = None
    community_rank_3: Optional[int] = None
    counselling_college_3: Optional[str] = None
    email: Optional[str]
    created_at: Optional[datetime]


class MessageResponse(BaseModel):
    message: str
    student: StudentResponse


def insert_student_data(student_data: StudentCreate, conn):
    """Insert student data into all related tables"""
    cursor = conn.cursor()
    
    try:
        # 1. Insert into student table
        cursor.execute("""
            INSERT INTO student (
                student_id, batch_id, student_name, dob, grade, community,
                enrollment_year, course, board, gender, student_mobile,
                aadhar_no, apaar_id, email, photo_url, school_name
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING student_no, student_id, student_name, batch_id, grade, gender, email, created_at;
        """, (
            student_data.student_id, student_data.batch_id, student_data.student_name,
            student_data.dob, student_data.grade, student_data.community,
            student_data.enrollment_year, student_data.course, student_data.board,
            student_data.gender, student_data.student_mobile, student_data.aadhar_no,
            student_data.apaar_id, student_data.email, student_data.photo_url,
            student_data.school_name
        ))
        
        student_result = cursor.fetchone()
        student_no = student_result[0]
        
        # 2. Insert into parent_info table
        cursor.execute("""
            INSERT INTO parent_info (
                student_no, guardian_name, father_name, mother_name, sibling_name,
                guardian_occupation, father_occupation, mother_occupation, sibling_grade,
                guardian_mobile, mother_mobile, father_mobile,
                sibling_school, sibling_college,
                guardian_email, mother_email, father_email
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
        """, (
            student_no, student_data.guardian_name, student_data.father_name,
            student_data.mother_name, student_data.sibling_name,
            student_data.guardian_occupation, student_data.father_occupation,
            student_data.mother_occupation, student_data.sibling_grade,
            student_data.guardian_mobile, student_data.mother_mobile,
            student_data.father_mobile, student_data.sibling_school,
            student_data.sibling_college, student_data.guardian_email,
            student_data.mother_email, student_data.father_email
        ))
        
        # 3. Insert into tenth_mark table (if data provided)
        if student_data.tenth_school_name or student_data.tenth_year_of_passing:
            cursor.execute("""
                INSERT INTO tenth_mark (
                    student_no, school_name, year_of_passing, board_of_study,
                    english, tamil, hindi, maths, science, social_science, total_marks
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
            """, (
                student_no, student_data.tenth_school_name,
                student_data.tenth_year_of_passing, student_data.tenth_board_of_study,
                student_data.tenth_english, student_data.tenth_tamil,
                student_data.tenth_hindi, student_data.tenth_maths,
                student_data.tenth_science, student_data.tenth_social_science,
                student_data.tenth_total_marks
            ))
        
        # 4. Insert into twelfth_mark table (if data provided)
        if student_data.twelfth_school_name or student_data.twelfth_year_of_passing:
            cursor.execute("""
                INSERT INTO twelfth_mark (
                    student_no, school_name, year_of_passing, board_of_study,
                    english, physics, maths, chemistry, biology, computer_science, tamil, total_marks
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
            """, (
                student_no, student_data.twelfth_school_name,
                student_data.twelfth_year_of_passing, student_data.twelfth_board_of_study,
                student_data.twelfth_english, student_data.twelfth_physics,
                student_data.twelfth_maths, student_data.twelfth_chemistry,
                student_data.twelfth_biology, student_data.twelfth_computer_science,
                student_data.twelfth_tamil, student_data.twelfth_total_marks
            ))
        
        # 5. Insert entrance exams (if provided)
        if (student_data.entrance_exam_1 or student_data.entrance_exam_2 or student_data.entrance_exam_3):
            cursor.execute("""
                INSERT INTO entrance_exams (
                    student_no, 
                    entrance_exam_1, entrance_exam_1_percentile, entrance_exam_1_mark,
                    entrance_exam_2, entrance_exam_2_percentile, entrance_exam_2_mark,
                    entrance_exam_3, entrance_exam_3_percentile, entrance_exam_3_mark
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
            """, (
                student_no, 
                student_data.entrance_exam_1, student_data.entrance_exam_1_percentile, student_data.entrance_exam_1_mark,
                student_data.entrance_exam_2, student_data.entrance_exam_2_percentile, student_data.entrance_exam_2_mark,
                student_data.entrance_exam_3, student_data.entrance_exam_3_percentile, student_data.entrance_exam_3_mark
            ))
        
        # 6. Insert counselling details (if provided)
        if (student_data.counselling_forum_1 or student_data.counselling_college_1 or
            student_data.counselling_forum_2 or student_data.counselling_college_2 or
            student_data.counselling_forum_3 or student_data.counselling_college_3):
            cursor.execute("""
                INSERT INTO counselling_detail (
                    student_no, 
                    counselling_forum_1, counselling_round_1, all_india_rank_1, community_rank_1, counselling_college_1,
                    counselling_forum_2, counselling_round_2, all_india_rank_2, community_rank_2, counselling_college_2,
                    counselling_forum_3, counselling_round_3, all_india_rank_3, community_rank_3, counselling_college_3
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
            """, (
                student_no, 
                student_data.counselling_forum_1, student_data.counselling_round_1, student_data.all_india_rank_1, student_data.community_rank_1, student_data.counselling_college_1,
                student_data.counselling_forum_2, student_data.counselling_round_2, student_data.all_india_rank_2, student_data.community_rank_2, student_data.counselling_college_2,
                student_data.counselling_forum_3, student_data.counselling_round_3, student_data.all_india_rank_3, student_data.community_rank_3, student_data.counselling_college_3
            ))
        
        return student_result
        
    except Exception as e:
        raise e


@app.post("/api/student", response_model=MessageResponse, status_code=status.HTTP_201_CREATED)
async def create_student(student: StudentCreate, current_user: dict = Depends(get_current_user)):
    """
    Create a new student with all related information
    """
    conn = None
    try:
        # Connect to database
        conn = get_db_connection()
        if not conn:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Database connection failed"
            )
        
        # Check if batch exists
        cursor = conn.cursor()
        cursor.execute("SELECT batch_id FROM batch WHERE batch_id = %s", (student.batch_id,))
        if not cursor.fetchone():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Batch with ID {student.batch_id} not found"
            )
        
        # Insert student data
        result = insert_student_data(student, conn)
        
        # Commit the transaction
        conn.commit()
        
        # Prepare response
        student_response = StudentResponse(
            student_id=result[1],
            student_name=result[2],
            batch_id=result[3],
            grade=result[4],
            gender=result[5],
            email=result[6],
            created_at=result[7]
        )
        
        cursor.close()
        conn.close()
        
        return MessageResponse(
            message="Student added successfully",
            student=student_response
        )
        
    except psycopg2.IntegrityError as e:
        if conn:
            conn.rollback()
            conn.close()
        error_msg = str(e)
        if (
            "student_pkey" in error_msg
            or "student_student_id_key" in error_msg
            or "unique constraint" in error_msg.lower()
        ):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Student with ID {student.student_id} already exists"
            )
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Database integrity error: {str(e)}"
        )
    
    except psycopg2.Error as e:
        if conn:
            conn.rollback()
            conn.close()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Database error: {str(e)}"
        )
    
    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Server error: {str(e)}"
        )


@app.post("/api/student/upload", status_code=status.HTTP_201_CREATED)
async def upload_students_excel(
    file: UploadFile = File(...),
    batch_id: int = Form(...),
    current_user: dict = Depends(get_current_user)
):
    """
    Upload students from Excel file (.xlsx)
    
    The Excel file should have the following columns:
    - Required: student_id, student_name
    - All other fields are optional
    """
    conn = None
    try:
        # Validate file type
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File must be an Excel file (.xlsx or .xls)"
            )
        
        # Read Excel file
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        # Replace NaN with None
        df = df.where(pd.notna(df), None)
        
        # Validate required columns
        if 'student_id' not in df.columns or 'student_name' not in df.columns:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel file must contain 'student_id' and 'student_name' columns"
            )
        
        # Connect to database
        conn = get_db_connection()
        if not conn:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Database connection failed"
            )
        
        # Check if batch exists
        cursor = conn.cursor()
        cursor.execute("SELECT batch_id FROM batch WHERE batch_id = %s", (batch_id,))
        if not cursor.fetchone():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Batch with ID {batch_id} not found"
            )
        
        success_count = 0
        error_count = 0
        errors = []
        
        # Process each row
        for index, row in df.iterrows():
            try:
                # Create SAVEPOINT at the start so it exists for rollback on ANY error
                cursor.execute("SAVEPOINT student_row")
                
                # Helper to safely get a value from the row
                def get_val(col):
                    """Return value if column exists and is not NaN/None, else None"""
                    if col in row and pd.notna(row[col]):
                        return row[col]
                    return None
                
                # Convert date string to date object if present
                dob_value = None
                if get_val('dob') is not None:
                    try:
                        dob_value = pd.to_datetime(row['dob']).date()
                    except:
                        dob_value = None
                

                
                # Create student object with safe type conversions
                student_data = StudentCreate(
                    student_id=safe_str(row['student_id']),
                    batch_id=batch_id,
                    student_name=safe_str(row['student_name']),
                    dob=dob_value,
                    grade=safe_str(get_val('grade')),
                    community=safe_str(get_val('community')),
                    enrollment_year=safe_int(get_val('enrollment_year')),
                    course=safe_str(get_val('course')),
                    board=safe_str(get_val('board')),
                    gender=safe_str(get_val('gender')),
                    student_mobile=safe_str(get_val('student_mobile')),
                    aadhar_no=safe_str(get_val('aadhar_no')),
                    apaar_id=safe_str(get_val('apaar_id')),
                    email=safe_str(get_val('email')),
                    school_name=safe_str(get_val('school_name')),
                    
                    # Parent info
                    guardian_name=safe_str(get_val('guardian_name')),
                    guardian_occupation=safe_str(get_val('guardian_occupation')),
                    guardian_mobile=safe_str(get_val('guardian_mobile')),
                    guardian_email=safe_str(get_val('guardian_email')),
                    father_name=safe_str(get_val('father_name')),
                    father_occupation=safe_str(get_val('father_occupation')),
                    father_mobile=safe_str(get_val('father_mobile')),
                    father_email=safe_str(get_val('father_email')),
                    mother_name=safe_str(get_val('mother_name')),
                    mother_occupation=safe_str(get_val('mother_occupation')),
                    mother_mobile=safe_str(get_val('mother_mobile')),
                    mother_email=safe_str(get_val('mother_email')),
                    sibling_name=safe_str(get_val('sibling_name')),
                    sibling_grade=safe_str(get_val('sibling_grade')),
                    sibling_school=safe_str(get_val('sibling_school')),
                    sibling_college=safe_str(get_val('sibling_college')),
                    
                    # 10th marks
                    tenth_school_name=safe_str(get_val('tenth_school_name')),
                    tenth_year_of_passing=safe_int(get_val('tenth_year_of_passing')),
                    tenth_board_of_study=safe_str(get_val('tenth_board_of_study')),
                    tenth_english=safe_int(get_val('tenth_english')),
                    tenth_tamil=safe_int(get_val('tenth_tamil')),
                    tenth_hindi=safe_int(get_val('tenth_hindi')),
                    tenth_maths=safe_int(get_val('tenth_maths')),
                    tenth_science=safe_int(get_val('tenth_science')),
                    tenth_social_science=safe_int(get_val('tenth_social_science')),
                    tenth_total_marks=safe_int(get_val('tenth_total_marks')),
                    
                    # 12th marks
                    twelfth_school_name=safe_str(get_val('twelfth_school_name')),
                    twelfth_year_of_passing=safe_int(get_val('twelfth_year_of_passing')),
                    twelfth_board_of_study=safe_str(get_val('twelfth_board_of_study')),
                    twelfth_english=safe_int(get_val('twelfth_english')),
                    twelfth_tamil=safe_int(get_val('twelfth_tamil')),
                    twelfth_physics=safe_int(get_val('twelfth_physics')),
                    twelfth_chemistry=safe_int(get_val('twelfth_chemistry')),
                    twelfth_maths=safe_int(get_val('twelfth_maths')),
                    twelfth_biology=safe_int(get_val('twelfth_biology')),
                    twelfth_computer_science=safe_int(get_val('twelfth_computer_science')),
                    twelfth_total_marks=safe_int(get_val('twelfth_total_marks')),
                    
                    # Entrance exams
                    entrance_exam_1=safe_str(get_val('entrance_exam_1')),
                    entrance_exam_1_percentile=safe_float(get_val('entrance_exam_1_percentile')),
                    entrance_exam_1_mark=safe_int(get_val('entrance_exam_1_mark')),
                    entrance_exam_2=safe_str(get_val('entrance_exam_2')),
                    entrance_exam_2_percentile=safe_float(get_val('entrance_exam_2_percentile')),
                    entrance_exam_2_mark=safe_int(get_val('entrance_exam_2_mark')),
                    entrance_exam_3=safe_str(get_val('entrance_exam_3')),
                    entrance_exam_3_percentile=safe_float(get_val('entrance_exam_3_percentile')),
                    entrance_exam_3_mark=safe_int(get_val('entrance_exam_3_mark')),
                    
                    # Counselling
                    counselling_forum_1=safe_str(get_val('counselling_forum_1')),
                    counselling_round_1=safe_int(get_val('counselling_round_1')),
                    all_india_rank_1=safe_int(get_val('all_india_rank_1')),
                    community_rank_1=safe_int(get_val('community_rank_1')),
                    counselling_college_1=safe_str(get_val('counselling_college_1')),
                    counselling_forum_2=safe_str(get_val('counselling_forum_2')),
                    counselling_round_2=safe_int(get_val('counselling_round_2')),
                    all_india_rank_2=safe_int(get_val('all_india_rank_2')),
                    community_rank_2=safe_int(get_val('community_rank_2')),
                    counselling_college_2=safe_str(get_val('counselling_college_2')),
                    counselling_forum_3=safe_str(get_val('counselling_forum_3')),
                    counselling_round_3=safe_int(get_val('counselling_round_3')),
                    all_india_rank_3=safe_int(get_val('all_india_rank_3')),
                    community_rank_3=safe_int(get_val('community_rank_3')),
                    counselling_college_3=safe_str(get_val('counselling_college_3'))
                )
                
                # Use SAVEPOINT so a single row failure doesn't abort the whole transaction
                insert_student_data(student_data, conn)
                cursor.execute("RELEASE SAVEPOINT student_row")
                success_count += 1
                
            except Exception as e:
                # Rollback only this row's changes, keep the transaction alive
                import traceback
                traceback.print_exc()
                cursor.execute("ROLLBACK TO SAVEPOINT student_row")
                error_count += 1
                error_msg = str(e)
                if (
                    "student_pkey" in error_msg
                    or "student_student_id_key" in error_msg
                    or "unique constraint" in error_msg.lower()
                    or "already exists" in error_msg.lower()
                ):
                    error_msg = f"Student {row.get('student_id', 'N/A')} already exists in the database"
                errors.append({
                    'row': int(index) + 2,  # +2 because Excel is 1-indexed and has header
                    'student_id': safe_str(row.get('student_id', 'N/A')),
                    'error': error_msg
                })
        
        # Commit all successful insertions
        conn.commit()
        cursor.close()
        conn.close()
        
        if success_count > 0 and error_count == 0:
            message = "All students uploaded successfully"
        elif success_count > 0:
            message = f"Upload partially completed — {error_count} student(s) failed"
        else:
            message = "Upload failed — no students were added to the database"
        
        return {
            "message": message,
            "success_count": success_count,
            "error_count": error_count,
            "errors": errors[:50] if errors else []
        }
        
    except HTTPException:
        raise
    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error processing file: {str(e)}"
        )


@app.get("/api/student/batch/{batch_id}")
async def get_students_by_batch(batch_id: int, current_user: dict = Depends(get_current_user)):
    """
    Get all students in a specific batch with their basic information
    """
    conn = None
    try:
        conn = get_db_connection()
        if not conn:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Database connection failed"
            )
        
        cursor = conn.cursor()
        
        # Fetch students with basic info
        query = """
            SELECT 
                s.student_no,
                s.student_id,
                s.student_name,
                s.gender,
                s.dob,
                s.community,
                s.grade,
                s.enrollment_year,
                s.course,
                s.board,
                s.student_mobile,
                s.email,
                s.created_at
            FROM student s
            WHERE s.batch_id = %s
            ORDER BY
                CASE WHEN s.student_id ~ '^[0-9]+$' THEN 0 ELSE 1 END,
                CASE WHEN s.student_id ~ '^[0-9]+$' THEN s.student_id::BIGINT END,
                s.student_id ASC,
                s.student_name ASC
        """
        
        cursor.execute(query, (batch_id,))
        rows = cursor.fetchall()
        
        students = []
        for row in rows:
            student = {
                "student_no": row[0],
                "student_id": row[1],
                "student_name": row[2],
                "gender": row[3],
                "dob": row[4].isoformat() if row[4] else None,
                "community": row[5],
                "grade": row[6],
                "enrollment_year": row[7],
                "course": row[8],
                "board": row[9],
                "student_mobile": row[10],
                "email": row[11],
                "created_at": row[12].isoformat() if row[12] else None
            }
            students.append(student)
        
        cursor.close()
        conn.close()
        
        return {
            "batch_id": batch_id,
            "count": len(students),
            "students": students
        }
        
    except HTTPException:
        raise
    except Exception as e:
        if conn:
            conn.close()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching students: {str(e)}"
        )


@app.get("/api/student/{student_no}")
async def get_student_details(student_no: int, current_user: dict = Depends(get_current_user)):
    """
    Get complete student details from all related tables
    """
    conn = None
    try:
        conn = get_db_connection()
        if not conn:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Database connection failed"
            )
        
        cursor = conn.cursor()
        
        # Fetch student basic info
        cursor.execute("""
            SELECT student_no, student_id, batch_id, student_name, dob, grade, community, 
                   enrollment_year, course, board, gender, student_mobile, 
                   aadhar_no, apaar_id, email, school_name, created_at, photo_url
            FROM student WHERE student_no = %s
        """, (student_no,))
        
        student_row = cursor.fetchone()
        if not student_row:
            cursor.close()
            conn.close()
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Student {student_no} not found"
            )
        
        student_data = {
            "student_no": student_row[0],
            "student_id": student_row[1],
            "batch_id": student_row[2],
            "student_name": student_row[3],
            "dob": student_row[4].isoformat() if student_row[4] else None,
            "grade": student_row[5],
            "community": student_row[6],
            "enrollment_year": student_row[7],
            "course": student_row[8],
            "board": student_row[9],
            "gender": student_row[10],
            "student_mobile": student_row[11],
            "aadhar_no": student_row[12],
            "apaar_id": student_row[13],
            "email": student_row[14],
            "school_name": student_row[15],
            "created_at": student_row[16].isoformat() if student_row[16] else None,
            "photo_url": student_row[17]
        }
        
        # Fetch parent info
        cursor.execute("""
            SELECT guardian_name, guardian_occupation, guardian_mobile, guardian_email,
                   father_name, father_occupation, father_mobile, father_email,
                   mother_name, mother_occupation, mother_mobile, mother_email,
                   sibling_name, sibling_grade, sibling_school, sibling_college
            FROM parent_info WHERE student_no = %s
        """, (student_no,))
        
        parent_row = cursor.fetchone()
        if parent_row:
            student_data.update({
                "guardian_name": parent_row[0],
                "guardian_occupation": parent_row[1],
                "guardian_mobile": parent_row[2],
                "guardian_email": parent_row[3],
                "father_name": parent_row[4],
                "father_occupation": parent_row[5],
                "father_mobile": parent_row[6],
                "father_email": parent_row[7],
                "mother_name": parent_row[8],
                "mother_occupation": parent_row[9],
                "mother_mobile": parent_row[10],
                "mother_email": parent_row[11],
                "sibling_name": parent_row[12],
                "sibling_grade": parent_row[13],
                "sibling_school": parent_row[14],
                "sibling_college": parent_row[15]
            })
        
        # Fetch 10th marks
        cursor.execute("""
            SELECT school_name, year_of_passing, board_of_study, english, tamil, hindi,
                   maths, science, social_science, total_marks
            FROM tenth_mark WHERE student_no = %s
        """, (student_no,))
        
        tenth_row = cursor.fetchone()
        if tenth_row:
            student_data.update({
                "tenth_school_name": tenth_row[0],
                "tenth_year_of_passing": tenth_row[1],
                "tenth_board_of_study": tenth_row[2],
                "tenth_english": tenth_row[3],
                "tenth_tamil": tenth_row[4],
                "tenth_hindi": tenth_row[5],
                "tenth_maths": tenth_row[6],
                "tenth_science": tenth_row[7],
                "tenth_social_science": tenth_row[8],
                "tenth_total_marks": tenth_row[9]
            })
        
        # Fetch 12th marks
        cursor.execute("""
            SELECT school_name, year_of_passing, board_of_study, english, physics,
                   maths, chemistry, biology, computer_science, tamil, total_marks
            FROM twelfth_mark WHERE student_no = %s
        """, (student_no,))
        
        twelfth_row = cursor.fetchone()
        if twelfth_row:
            student_data.update({
                "twelfth_school_name": twelfth_row[0],
                "twelfth_year_of_passing": twelfth_row[1],
                "twelfth_board_of_study": twelfth_row[2],
                "twelfth_english": twelfth_row[3],
                "twelfth_physics": twelfth_row[4],
                "twelfth_maths": twelfth_row[5],
                "twelfth_chemistry": twelfth_row[6],
                "twelfth_biology": twelfth_row[7],
                "twelfth_computer_science": twelfth_row[8],
                "twelfth_tamil": twelfth_row[9],
                "twelfth_total_marks": twelfth_row[10]
            })
        
        # Fetch entrance exams
        cursor.execute("""
            SELECT entrance_exam_1, entrance_exam_1_percentile, entrance_exam_1_mark,
                   entrance_exam_2, entrance_exam_2_percentile, entrance_exam_2_mark,
                   entrance_exam_3, entrance_exam_3_percentile, entrance_exam_3_mark
            FROM entrance_exams WHERE student_no = %s
        """, (student_no,))
        
        entrance_row = cursor.fetchone()
        if entrance_row:
            student_data.update({
                "entrance_exam_1": entrance_row[0],
                "entrance_exam_1_percentile": float(entrance_row[1]) if entrance_row[1] is not None else None,
                "entrance_exam_1_mark": entrance_row[2],
                "entrance_exam_2": entrance_row[3],
                "entrance_exam_2_percentile": float(entrance_row[4]) if entrance_row[4] is not None else None,
                "entrance_exam_2_mark": entrance_row[5],
                "entrance_exam_3": entrance_row[6],
                "entrance_exam_3_percentile": float(entrance_row[7]) if entrance_row[7] is not None else None,
                "entrance_exam_3_mark": entrance_row[8]
            })
        
        # Fetch counselling details
        cursor.execute("""
            SELECT counselling_forum_1, counselling_round_1, all_india_rank_1, community_rank_1, counselling_college_1,
                   counselling_forum_2, counselling_round_2, all_india_rank_2, community_rank_2, counselling_college_2,
                   counselling_forum_3, counselling_round_3, all_india_rank_3, community_rank_3, counselling_college_3
            FROM counselling_detail WHERE student_no = %s
        """, (student_no,))
        
        counselling_row = cursor.fetchone()
        if counselling_row:
            student_data.update({
                "counselling_forum_1": counselling_row[0],
                "counselling_round_1": counselling_row[1],
                "all_india_rank_1": counselling_row[2],
                "community_rank_1": counselling_row[3],
                "counselling_college_1": counselling_row[4],
                "counselling_forum_2": counselling_row[5],
                "counselling_round_2": counselling_row[6],
                "all_india_rank_2": counselling_row[7],
                "community_rank_2": counselling_row[8],
                "counselling_college_2": counselling_row[9],
                "counselling_forum_3": counselling_row[10],
                "counselling_round_3": counselling_row[11],
                "all_india_rank_3": counselling_row[12],
                "community_rank_3": counselling_row[13],
                "counselling_college_3": counselling_row[14]
            })
        
        cursor.close()
        conn.close()
        
        return student_data
        
    except HTTPException:
        raise
    except Exception as e:
        if conn:
            conn.close()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching student details: {str(e)}"
        )


@app.put("/api/student/{student_no}")
async def update_student(student_no: int, updates: StudentUpdate, current_user: dict = Depends(get_current_user)):
    """
    Update student details - only updates fields that are provided (partial update)
    """
    conn = None
    try:
        conn = get_db_connection()
        if not conn:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Database connection failed"
            )
        
        cursor = conn.cursor()
        
        # Check if student exists
        cursor.execute("SELECT student_id FROM student WHERE student_no = %s", (student_no,))
        existing_student = cursor.fetchone()
        if not existing_student:
            cursor.close()
            conn.close()
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Student {student_no} not found"
            )
        student_id = existing_student[0]
        
        # Build dynamic UPDATE query for student table - only update provided fields
        student_fields = {
            'student_name': updates.student_name,
            'dob': updates.dob,
            'grade': updates.grade,
            'community': updates.community,
            'enrollment_year': updates.enrollment_year,
            'course': updates.course,
            'board': updates.board,
            'gender': updates.gender,
            'student_mobile': updates.student_mobile,
            'aadhar_no': updates.aadhar_no,
            'apaar_id': updates.apaar_id,
            'email': updates.email,
            'school_name': updates.school_name
        }
        
        # Filter out None values (fields not provided)
        student_updates = {k: v for k, v in student_fields.items() if v is not None}
        
        if student_updates:
            set_clause = ", ".join([f"{k} = %s" for k in student_updates.keys()])
            query = f"UPDATE student SET {set_clause} WHERE student_no = %s"
            cursor.execute(query, list(student_updates.values()) + [student_no])
        
        # Update parent_info table (UPSERT logic)
        parent_fields = {
            'guardian_name': updates.guardian_name,
            'guardian_occupation': updates.guardian_occupation,
            'guardian_mobile': updates.guardian_mobile,
            'guardian_email': updates.guardian_email,
            'father_name': updates.father_name,
            'father_occupation': updates.father_occupation,
            'father_mobile': updates.father_mobile,
            'father_email': updates.father_email,
            'mother_name': updates.mother_name,
            'mother_occupation': updates.mother_occupation,
            'mother_mobile': updates.mother_mobile,
            'mother_email': updates.mother_email,
            'sibling_name': updates.sibling_name,
            'sibling_grade': updates.sibling_grade,
            'sibling_school': updates.sibling_school,
            'sibling_college': updates.sibling_college
        }
        
        parent_updates = {k: v for k, v in parent_fields.items() if v is not None}
        
        if parent_updates:
            # Check if parent_info exists
            cursor.execute("SELECT student_no FROM parent_info WHERE student_no = %s", (student_no,))
            parent_exists = cursor.fetchone()
            
            if parent_exists:
                set_clause = ", ".join([f"{k} = %s" for k in parent_updates.keys()])
                query = f"UPDATE parent_info SET {set_clause} WHERE student_no = %s"
                cursor.execute(query, list(parent_updates.values()) + [student_no])
            else:
                columns = ['student_no'] + list(parent_updates.keys())
                placeholders = ', '.join(['%s'] * len(columns))
                query = f"INSERT INTO parent_info ({', '.join(columns)}) VALUES ({placeholders})"
                cursor.execute(query, [student_no] + list(parent_updates.values()))
        
        # Update tenth_mark table (UPSERT logic)
        tenth_fields = {
            'school_name': updates.tenth_school_name,
            'year_of_passing': updates.tenth_year_of_passing,
            'board_of_study': updates.tenth_board_of_study,
            'english': updates.tenth_english,
            'tamil': updates.tenth_tamil,
            'hindi': updates.tenth_hindi,
            'maths': updates.tenth_maths,
            'science': updates.tenth_science,
            'social_science': updates.tenth_social_science,
            'total_marks': updates.tenth_total_marks
        }
        
        tenth_updates = {k: v for k, v in tenth_fields.items() if v is not None}
        
        if tenth_updates:
            cursor.execute("SELECT student_no FROM tenth_mark WHERE student_no = %s", (student_no,))
            tenth_exists = cursor.fetchone()
            
            if tenth_exists:
                set_clause = ", ".join([f"{k} = %s" for k in tenth_updates.keys()])
                query = f"UPDATE tenth_mark SET {set_clause} WHERE student_no = %s"
                cursor.execute(query, list(tenth_updates.values()) + [student_no])
            else:
                columns = ['student_no'] + list(tenth_updates.keys())
                placeholders = ', '.join(['%s'] * len(columns))
                query = f"INSERT INTO tenth_mark ({', '.join(columns)}) VALUES ({placeholders})"
                cursor.execute(query, [student_no] + list(tenth_updates.values()))
        
        # Update twelfth_mark table (UPSERT logic)
        twelfth_fields = {
            'school_name': updates.twelfth_school_name,
            'year_of_passing': updates.twelfth_year_of_passing,
            'board_of_study': updates.twelfth_board_of_study,
            'english': updates.twelfth_english,
            'tamil': updates.twelfth_tamil,
            'physics': updates.twelfth_physics,
            'chemistry': updates.twelfth_chemistry,
            'maths': updates.twelfth_maths,
            'biology': updates.twelfth_biology,
            'computer_science': updates.twelfth_computer_science,
            'total_marks': updates.twelfth_total_marks
        }
        
        twelfth_updates = {k: v for k, v in twelfth_fields.items() if v is not None}
        
        if twelfth_updates:
            cursor.execute("SELECT student_no FROM twelfth_mark WHERE student_no = %s", (student_no,))
            twelfth_exists = cursor.fetchone()
            
            if twelfth_exists:
                set_clause = ", ".join([f"{k} = %s" for k in twelfth_updates.keys()])
                query = f"UPDATE twelfth_mark SET {set_clause} WHERE student_no = %s"
                cursor.execute(query, list(twelfth_updates.values()) + [student_no])
            else:
                columns = ['student_no'] + list(twelfth_updates.keys())
                placeholders = ', '.join(['%s'] * len(columns))
                query = f"INSERT INTO twelfth_mark ({', '.join(columns)}) VALUES ({placeholders})"
                cursor.execute(query, [student_no] + list(twelfth_updates.values()))
        
        # Update entrance_exams table (UPSERT logic)
        entrance_fields = {
            'entrance_exam_1': updates.entrance_exam_1,
            'entrance_exam_1_percentile': updates.entrance_exam_1_percentile,
            'entrance_exam_1_mark': updates.entrance_exam_1_mark,
            'entrance_exam_2': updates.entrance_exam_2,
            'entrance_exam_2_percentile': updates.entrance_exam_2_percentile,
            'entrance_exam_2_mark': updates.entrance_exam_2_mark,
            'entrance_exam_3': updates.entrance_exam_3,
            'entrance_exam_3_percentile': updates.entrance_exam_3_percentile,
            'entrance_exam_3_mark': updates.entrance_exam_3_mark
        }
        
        entrance_updates = {k: v for k, v in entrance_fields.items() if v is not None}
        
        if entrance_updates:
            cursor.execute("SELECT student_no FROM entrance_exams WHERE student_no = %s", (student_no,))
            entrance_exists = cursor.fetchone()
            
            if entrance_exists:
                set_clause = ", ".join([f"{k} = %s" for k in entrance_updates.keys()])
                query = f"UPDATE entrance_exams SET {set_clause} WHERE student_no = %s"
                cursor.execute(query, list(entrance_updates.values()) + [student_no])
            else:
                columns = ['student_no'] + list(entrance_updates.keys())
                placeholders = ', '.join(['%s'] * len(columns))
                query = f"INSERT INTO entrance_exams ({', '.join(columns)}) VALUES ({placeholders})"
                cursor.execute(query, [student_no] + list(entrance_updates.values()))
                
        # Update counselling_detail table (UPSERT logic)
        counselling_fields = {
            'forum': updates.counselling_forum,
            'round': updates.counselling_round,
            'college_alloted': updates.counselling_college_alloted,
            'year_of_completion': updates.counselling_year_of_completion
        }
        
        counselling_updates = {k: v for k, v in counselling_fields.items() if v is not None}
        
        if counselling_updates:
            cursor.execute("SELECT student_no FROM counselling_detail WHERE student_no = %s", (student_no,))
            counselling_exists = cursor.fetchone()
            
            if counselling_exists:
                set_clause = ", ".join([f"{k} = %s" for k in counselling_updates.keys()])
                query = f"UPDATE counselling_detail SET {set_clause} WHERE student_no = %s"
                cursor.execute(query, list(counselling_updates.values()) + [student_no])
            else:
                columns = ['student_no'] + list(counselling_updates.keys())
                placeholders = ', '.join(['%s'] * len(columns))
                query = f"INSERT INTO counselling_detail ({', '.join(columns)}) VALUES ({placeholders})"
                cursor.execute(query, [student_no] + list(counselling_updates.values()))
        
        # Commit all changes
        conn.commit()
        cursor.close()
        conn.close()
        
        return {
            "message": f"Student {student_no} updated successfully",
            "student_no": student_no,
            "student_id": student_id,
            "updated_fields": len(student_updates) + len(parent_updates) + len(tenth_updates) + len(twelfth_updates) + len(counselling_updates)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error updating student: {str(e)}"
        )


# ==================== BULK EDIT VIA EXCEL ====================

# Column header -> (table, db_column, type)
# This maps every possible Excel header to its table and column for update.
BULK_EDIT_FIELD_MAP = {
    # student table
    "student_name":       ("student", "student_name", "str"),
    "dob":                ("student", "dob", "date"),
    "grade":              ("student", "grade", "str"),
    "community":          ("student", "community", "str"),
    "enrollment_year":    ("student", "enrollment_year", "int"),
    "course":             ("student", "course", "str"),
    "board":             ("student", "board", "str"),
    "gender":             ("student", "gender", "str"),
    "student_mobile":     ("student", "student_mobile", "str"),
    "aadhar_no":          ("student", "aadhar_no", "str"),
    "apaar_id":           ("student", "apaar_id", "str"),
    "email":              ("student", "email", "str"),
    "school_name":        ("student", "school_name", "str"),
    # parent_info table
    "guardian_name":       ("parent_info", "guardian_name", "str"),
    "guardian_occupation": ("parent_info", "guardian_occupation", "str"),
    "guardian_mobile":     ("parent_info", "guardian_mobile", "str"),
    "guardian_email":      ("parent_info", "guardian_email", "str"),
    "father_name":         ("parent_info", "father_name", "str"),
    "father_occupation":   ("parent_info", "father_occupation", "str"),
    "father_mobile":       ("parent_info", "father_mobile", "str"),
    "father_email":        ("parent_info", "father_email", "str"),
    "mother_name":         ("parent_info", "mother_name", "str"),
    "mother_occupation":   ("parent_info", "mother_occupation", "str"),
    "mother_mobile":       ("parent_info", "mother_mobile", "str"),
    "mother_email":        ("parent_info", "mother_email", "str"),
    "sibling_name":        ("parent_info", "sibling_name", "str"),
    "sibling_grade":       ("parent_info", "sibling_grade", "str"),
    "sibling_school":      ("parent_info", "sibling_school", "str"),
    "sibling_college":     ("parent_info", "sibling_college", "str"),
    # tenth_mark table
    "tenth_school_name":    ("tenth_mark", "school_name", "str"),
    "tenth_year_of_passing":("tenth_mark", "year_of_passing", "int"),
    "tenth_board_of_study": ("tenth_mark", "board_of_study", "str"),
    "tenth_english":        ("tenth_mark", "english", "int"),
    "tenth_tamil":          ("tenth_mark", "tamil", "int"),
    "tenth_hindi":          ("tenth_mark", "hindi", "int"),
    "tenth_maths":          ("tenth_mark", "maths", "int"),
    "tenth_science":        ("tenth_mark", "science", "int"),
    "tenth_social_science": ("tenth_mark", "social_science", "int"),
    "tenth_total_marks":    ("tenth_mark", "total_marks", "int"),
    # twelfth_mark table
    "twelfth_school_name":    ("twelfth_mark", "school_name", "str"),
    "twelfth_year_of_passing":("twelfth_mark", "year_of_passing", "int"),
    "twelfth_board_of_study": ("twelfth_mark", "board_of_study", "str"),
    "twelfth_english":        ("twelfth_mark", "english", "int"),
    "twelfth_tamil":          ("twelfth_mark", "tamil", "int"),
    "twelfth_physics":        ("twelfth_mark", "physics", "int"),
    "twelfth_chemistry":      ("twelfth_mark", "chemistry", "int"),
    "twelfth_maths":          ("twelfth_mark", "maths", "int"),
    "twelfth_biology":        ("twelfth_mark", "biology", "int"),
    "twelfth_computer_science":("twelfth_mark", "computer_science", "int"),
    "twelfth_total_marks":    ("twelfth_mark", "total_marks", "int"),
    # entrance_exams table
    "entrance_exam_1":           ("entrance_exams", "entrance_exam_1", "str"),
    "entrance_exam_1_percentile":("entrance_exams", "entrance_exam_1_percentile", "float"),
    "entrance_exam_1_mark":      ("entrance_exams", "entrance_exam_1_mark", "int"),
    "entrance_exam_2":           ("entrance_exams", "entrance_exam_2", "str"),
    "entrance_exam_2_percentile":("entrance_exams", "entrance_exam_2_percentile", "float"),
    "entrance_exam_2_mark":      ("entrance_exams", "entrance_exam_2_mark", "int"),
    "entrance_exam_3":           ("entrance_exams", "entrance_exam_3", "str"),
    "entrance_exam_3_percentile":("entrance_exams", "entrance_exam_3_percentile", "float"),
    "entrance_exam_3_mark":      ("entrance_exams", "entrance_exam_3_mark", "int"),
    # counselling_detail table
    "counselling_forum_1":           ("counselling_detail", "counselling_forum_1", "str"),
    "counselling_round_1":           ("counselling_detail", "counselling_round_1", "int"),
    "all_india_rank_1":              ("counselling_detail", "all_india_rank_1", "int"),
    "community_rank_1":              ("counselling_detail", "community_rank_1", "int"),
    "counselling_college_1":         ("counselling_detail", "counselling_college_1", "str"),
    "counselling_forum_2":           ("counselling_detail", "counselling_forum_2", "str"),
    "counselling_round_2":           ("counselling_detail", "counselling_round_2", "int"),
    "all_india_rank_2":              ("counselling_detail", "all_india_rank_2", "int"),
    "community_rank_2":              ("counselling_detail", "community_rank_2", "int"),
    "counselling_college_2":         ("counselling_detail", "counselling_college_2", "str"),
    "counselling_forum_3":           ("counselling_detail", "counselling_forum_3", "str"),
    "counselling_round_3":           ("counselling_detail", "counselling_round_3", "int"),
    "all_india_rank_3":              ("counselling_detail", "all_india_rank_3", "int"),
    "community_rank_3":              ("counselling_detail", "community_rank_3", "int"),
    "counselling_college_3":         ("counselling_detail", "counselling_college_3", "str"),
}

# All possible Excel headers in display order
ALL_EDIT_COLUMNS = [
    "student_id", "student_name",
    # student info
    "dob", "grade", "community", "enrollment_year", "course", "board",
    "gender", "student_mobile", "aadhar_no", "apaar_id", "email", "school_name",
    # parent info
    "guardian_name", "guardian_occupation", "guardian_mobile", "guardian_email",
    "father_name", "father_occupation", "father_mobile", "father_email",
    "mother_name", "mother_occupation", "mother_mobile", "mother_email",
    "sibling_name", "sibling_grade", "sibling_school", "sibling_college",
    # 10th marks
    "tenth_school_name", "tenth_year_of_passing", "tenth_board_of_study",
    "tenth_english", "tenth_tamil", "tenth_hindi", "tenth_maths",
    "tenth_science", "tenth_social_science", "tenth_total_marks",
    # entrance exams
    "entrance_exam_1", "entrance_exam_1_percentile", "entrance_exam_1_mark",
    "entrance_exam_2", "entrance_exam_2_percentile", "entrance_exam_2_mark",
    "entrance_exam_3", "entrance_exam_3_percentile", "entrance_exam_3_mark",
    # 12th marks
    "twelfth_school_name", "twelfth_year_of_passing", "twelfth_board_of_study",
    "twelfth_english", "twelfth_tamil", "twelfth_physics", "twelfth_chemistry",
    "twelfth_maths", "twelfth_biology", "twelfth_computer_science", "twelfth_total_marks",
    # counselling
    "counselling_forum_1", "counselling_round_1", "all_india_rank_1", "community_rank_1", "counselling_college_1",
    "counselling_forum_2", "counselling_round_2", "all_india_rank_2", "community_rank_2", "counselling_college_2",
    "counselling_forum_3", "counselling_round_3", "all_india_rank_3", "community_rank_3", "counselling_college_3",
]


@app.get("/api/student/edit-template/{batch_id}")
async def download_edit_template(batch_id: int, current_user: dict = Depends(get_current_user)):
    """
    Generate and download an Excel file pre-filled with all existing student
    data for the given batch. Teachers edit cells they want to change and
    re-upload it.
    """
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Verify batch exists
        cursor.execute("SELECT batch_name FROM batch WHERE batch_id = %s", (batch_id,))
        batch_row = cursor.fetchone()
        if not batch_row:
            raise HTTPException(status_code=404, detail=f"Batch {batch_id} not found")
        batch_name = batch_row[0]

        # Fetch every student in the batch with ALL related data
        cursor.execute("""
            SELECT s.student_no, s.student_id
            FROM student s
            WHERE s.batch_id = %s
            ORDER BY
                CASE WHEN s.student_id ~ '^[0-9]+$' THEN 0 ELSE 1 END,
                CASE WHEN s.student_id ~ '^[0-9]+$' THEN s.student_id::BIGINT END,
                s.student_id ASC,
                s.student_name ASC
        """, (batch_id,))
        student_rows = cursor.fetchall()

        if not student_rows:
            raise HTTPException(status_code=404, detail="No students found in this batch")

        # For each student, pull complete data (reuse the single-student query pattern)
        rows_data = []
        for student_no, sid in student_rows:
            row = {}
            # student table
            cursor.execute("""
                SELECT student_id, student_name, dob, grade, community,
                       enrollment_year, course, branch, gender, student_mobile,
                       aadhar_no, apaar_id, email, school_name
                FROM student WHERE student_no = %s
            """, (student_no,))
            s = cursor.fetchone()
            if s:
                row["student_id"] = s[0]
                row["student_name"] = s[1]
                row["dob"] = s[2].isoformat() if s[2] else None
                row["grade"] = s[3]
                row["community"] = s[4]
                row["enrollment_year"] = s[5]
                row["course"] = s[6]
                row["board"] = s[7]
                row["gender"] = s[8]
                row["student_mobile"] = s[9]
                row["aadhar_no"] = s[10]
                row["apaar_id"] = s[11]
                row["email"] = s[12]
                row["school_name"] = s[13]

            # parent_info
            cursor.execute("""
                SELECT guardian_name, guardian_occupation, guardian_mobile, guardian_email,
                       father_name, father_occupation, father_mobile, father_email,
                       mother_name, mother_occupation, mother_mobile, mother_email,
                       sibling_name, sibling_grade, sibling_school, sibling_college
                FROM parent_info WHERE student_no = %s
            """, (student_no,))
            p = cursor.fetchone()
            if p:
                for i, key in enumerate([
                    "guardian_name", "guardian_occupation", "guardian_mobile", "guardian_email",
                    "father_name", "father_occupation", "father_mobile", "father_email",
                    "mother_name", "mother_occupation", "mother_mobile", "mother_email",
                    "sibling_name", "sibling_grade", "sibling_school", "sibling_college"
                ]):
                    row[key] = p[i]

            # tenth_mark
            cursor.execute("""
                SELECT school_name, year_of_passing, board_of_study,
                       english, tamil, hindi, maths, science, social_science, total_marks
                FROM tenth_mark WHERE student_no = %s
            """, (student_no,))
            t10 = cursor.fetchone()
            if t10:
                for i, key in enumerate([
                    "tenth_school_name", "tenth_year_of_passing", "tenth_board_of_study",
                    "tenth_english", "tenth_tamil", "tenth_hindi", "tenth_maths",
                    "tenth_science", "tenth_social_science", "tenth_total_marks"
                ]):
                    row[key] = t10[i]

            # twelfth_mark
            cursor.execute("""
                SELECT school_name, year_of_passing, board_of_study,
                       english, tamil, physics, chemistry, maths, biology,
                       computer_science, total_marks
                FROM twelfth_mark WHERE student_no = %s
            """, (student_no,))
            t12 = cursor.fetchone()
            if t12:
                for i, key in enumerate([
                    "twelfth_school_name", "twelfth_year_of_passing", "twelfth_board_of_study",
                    "twelfth_english", "twelfth_tamil", "twelfth_physics", "twelfth_chemistry",
                    "twelfth_maths", "twelfth_biology", "twelfth_computer_science", "twelfth_total_marks"
                ]):
                    row[key] = t12[i]

            # entrance_exams
            cursor.execute("""
                SELECT entrance_exam_1, entrance_exam_1_percentile, entrance_exam_1_mark,
                       entrance_exam_2, entrance_exam_2_percentile, entrance_exam_2_mark,
                       entrance_exam_3, entrance_exam_3_percentile, entrance_exam_3_mark
                FROM entrance_exams WHERE student_no = %s
            """, (student_no,))
            en = cursor.fetchone()
            if en:
                for i, key in enumerate([
                    "entrance_exam_1", "entrance_exam_1_percentile", "entrance_exam_1_mark",
                    "entrance_exam_2", "entrance_exam_2_percentile", "entrance_exam_2_mark",
                    "entrance_exam_3", "entrance_exam_3_percentile", "entrance_exam_3_mark"
                ]):
                    row[key] = en[i]

            # counselling_detail
            cursor.execute("""
                SELECT counselling_forum_1, counselling_round_1, all_india_rank_1, community_rank_1, counselling_college_1,
                       counselling_forum_2, counselling_round_2, all_india_rank_2, community_rank_2, counselling_college_2,
                       counselling_forum_3, counselling_round_3, all_india_rank_3, community_rank_3, counselling_college_3
                FROM counselling_detail WHERE student_no = %s
            """, (student_no,))
            co = cursor.fetchone()
            if co:
                for i, key in enumerate([
                    "counselling_forum_1", "counselling_round_1", "all_india_rank_1", "community_rank_1", "counselling_college_1",
                    "counselling_forum_2", "counselling_round_2", "all_india_rank_2", "community_rank_2", "counselling_college_2",
                    "counselling_forum_3", "counselling_round_3", "all_india_rank_3", "community_rank_3", "counselling_college_3"
                ]):
                    row[key] = co[i]

            rows_data.append(row)

        # Build Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Student Data"

        # Styling
        mandatory_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        optional_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        locked_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Write headers
        for col_idx, header in enumerate(ALL_EDIT_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if header in ("student_id", "student_name"):
                cell.fill = mandatory_fill
            else:
                cell.fill = optional_fill

        # Write data rows
        for row_idx, row_data in enumerate(rows_data, 2):
            for col_idx, header in enumerate(ALL_EDIT_COLUMNS, 1):
                value = row_data.get(header)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                # Lightly shade the student_id column so users know not to change it
                if header == "student_id":
                    cell.fill = locked_fill

        # Auto-width columns
        for col_idx, header in enumerate(ALL_EDIT_COLUMNS, 1):
            max_len = len(header) + 4
            for row_idx in range(2, len(rows_data) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)) + 2)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len, 35)

        # Instructions sheet
        ins_ws = wb.create_sheet("Instructions")
        instructions = [
            ["Bulk Edit Template — Instructions"],
            [""],
            ["1. The blue columns (student_id, student_name) are mandatory and identify each student."],
            ["2. Do NOT change the student_id column — it is used to match rows to the database."],
            ["3. Edit any green column to update that field. Leave cells blank to keep existing values."],
            ["4. You can DELETE entire columns you don't want to edit — they will be ignored."],
            ["5. Blank cells are skipped — existing data will NOT be erased."],
            ["6. Date format for 'dob': YYYY-MM-DD (e.g. 2005-06-15)"],
            ["7. Save and upload the file back on the Bulk Edit page."],
        ]
        for r, row in enumerate(instructions, 1):
            cell = ins_ws.cell(row=r, column=1, value=row[0])
            if r == 1:
                cell.font = Font(bold=True, size=14)
        ins_ws.column_dimensions['A'].width = 80

        # Return as downloadable file
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        safe_name = batch_name.replace(' ', '_') if batch_name else f'batch_{batch_id}'
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={safe_name}_Edit_Template.xlsx"}
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate edit template: {str(e)}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@app.post("/api/student/upload-update")
async def bulk_update_students(
    file: UploadFile = File(...),
    batch_id: int = Form(...),
    current_user: dict = Depends(get_current_user)
):
    """
    Upload an edited Excel file to bulk-update student records.
    Rules:
      - student_id and student_name columns are mandatory in the file.
      - student_id is used to identify which student to update.
      - student_name is required so rows can be validated.
      - All other columns are optional — only present, non-empty cells are updated.
      - Columns that don't exist in the file are completely ignored.
      - Blank cells are skipped (existing DB data preserved).
    """
    conn = None
    cursor = None
    try:
        # Read uploaded Excel
        contents = await file.read()
        try:
            df = pd.read_excel(io.BytesIO(contents), sheet_name=0, dtype=str)
        except Exception:
            raise HTTPException(status_code=400, detail="Could not read the Excel file. Please upload a valid .xlsx file.")

        # Normalize column names: strip, lowercase
        df.columns = [c.strip().lower() for c in df.columns]

        # Validate mandatory columns
        if "student_id" not in df.columns:
            raise HTTPException(status_code=400, detail="Missing mandatory column: 'student_id'")
        if "student_name" not in df.columns:
            raise HTTPException(status_code=400, detail="Missing mandatory column: 'student_name'")

        # Replace NaN with None
        df = df.replace({np.nan: None})

        conn = get_db_connection()
        cursor = conn.cursor()

        # Verify batch exists
        cursor.execute("SELECT batch_id FROM batch WHERE batch_id = %s", (batch_id,))
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail=f"Batch {batch_id} not found")

        # Figure out which editable columns are present in the uploaded file
        present_fields = {}
        for col in df.columns:
            if col in BULK_EDIT_FIELD_MAP:
                present_fields[col] = BULK_EDIT_FIELD_MAP[col]

        success_count = 0
        skipped_count = 0
        error_count = 0
        errors = []

        for idx, excel_row in df.iterrows():
            row_num = idx + 2  # Excel row (1-based header + 1)
            sid = excel_row.get("student_id")
            sname = excel_row.get("student_name")

            # Skip completely empty rows
            if not sid or str(sid).strip() == "":
                skipped_count += 1
                continue

            sid = str(sid).strip()
            # Remove trailing .0 from IDs read as floats
            if sid.endswith('.0') and sid[:-2].replace('-', '', 1).isdigit():
                sid = sid[:-2]

            # Validate student_name is present
            if not sname or str(sname).strip() == "":
                errors.append({"row": row_num, "student_id": sid, "error": "student_name is empty (mandatory)"})
                error_count += 1
                continue

            # Check student exists and belongs to this batch
            cursor.execute(
                """
                SELECT student_no
                FROM student
                WHERE student_id = %s AND batch_id = %s
                ORDER BY created_at DESC, student_no DESC
                LIMIT 1
                """,
                (sid, batch_id)
            )
            student_row = cursor.fetchone()
            if not student_row:
                errors.append({"row": row_num, "student_id": sid, "error": "Student not found in this batch"})
                error_count += 1
                continue
            student_no = student_row[0]

            try:
                # Group updates by table
                table_updates = {}  # table_name -> {db_col: value}
                for excel_col, (table, db_col, col_type) in present_fields.items():
                    raw = excel_row.get(excel_col)
                    if raw is None or str(raw).strip() == "":
                        continue  # skip blank cells

                    val = str(raw).strip()
                    # Remove trailing .0 from numeric strings
                    if val.endswith('.0') and val[:-2].replace('-', '', 1).isdigit():
                        val = val[:-2]

                    # Type coercion
                    if col_type == "int":
                        try:
                            val = int(float(val))
                        except (ValueError, TypeError):
                            errors.append({"row": row_num, "student_id": sid, "error": f"{excel_col}: expected a number, got '{val}'"})
                            continue
                    elif col_type == "float":
                        try:
                            val = float(val)
                        except (ValueError, TypeError):
                            errors.append({"row": row_num, "student_id": sid, "error": f"{excel_col}: expected a decimal number, got '{val}'"})
                            continue
                    elif col_type == "date":
                        try:
                            # Try multiple date formats
                            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
                                try:
                                    val = datetime.strptime(val, fmt).date()
                                    break
                                except ValueError:
                                    continue
                            else:
                                raise ValueError(f"Unrecognized date: {val}")
                        except Exception:
                            errors.append({"row": row_num, "student_id": sid, "error": f"{excel_col}: invalid date '{raw}'"})
                            continue

                    if table not in table_updates:
                        table_updates[table] = {}
                    table_updates[table][db_col] = val

                # Also update student_name in student table (it's mandatory and always present)
                name_val = str(sname).strip()
                if "student" not in table_updates:
                    table_updates["student"] = {}
                table_updates["student"]["student_name"] = name_val

                # Apply updates per table
                for table, fields in table_updates.items():
                    if not fields:
                        continue

                    if table == "student":
                        set_clause = ", ".join([f"{k} = %s" for k in fields.keys()])
                        query = f"UPDATE student SET {set_clause} WHERE student_no = %s"
                        cursor.execute(query, list(fields.values()) + [student_no])
                    else:
                        # For related tables use UPSERT: update if exists, insert if not
                        cursor.execute(
                            f"SELECT student_no FROM {table} WHERE student_no = %s", (student_no,)
                        )
                        exists = cursor.fetchone()

                        if exists:
                            set_clause = ", ".join([f"{k} = %s" for k in fields.keys()])
                            query = f"UPDATE {table} SET {set_clause} WHERE student_no = %s"
                            cursor.execute(query, list(fields.values()) + [student_no])
                        else:
                            columns = ["student_no"] + list(fields.keys())
                            placeholders = ", ".join(["%s"] * len(columns))
                            query = f"INSERT INTO {table} ({', '.join(columns)}) VALUES ({placeholders})"
                            cursor.execute(query, [student_no] + list(fields.values()))

                success_count += 1

            except Exception as row_err:
                errors.append({"row": row_num, "student_id": sid, "error": str(row_err)})
                error_count += 1

        conn.commit()

        return {
            "message": "Bulk update completed",
            "success_count": success_count,
            "skipped_count": skipped_count,
            "error_count": error_count,
            "errors": errors[:50]  # limit to first 50 errors
        }

    except HTTPException:
        raise
    except Exception as e:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"Bulk update failed: {str(e)}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@app.get("/api/student/template")
async def download_template(current_user: dict = Depends(get_current_user)):
    """
    Get the list of required columns for the Excel template
    """
    columns = {
        "required": ["student_id", "student_name"],
        "optional_student_info": [
            "dob", "grade", "community", "enrollment_year", "course", "board",
            "gender", "student_mobile", "aadhar_no", "apaar_id", "email", "school_name"
        ],
        "optional_parent_info": [
            "guardian_name", "guardian_occupation", "guardian_mobile", "guardian_email",
            "father_name", "father_occupation", "father_mobile", "father_email",
            "mother_name", "mother_occupation", "mother_mobile", "mother_email",
            "sibling_name", "sibling_grade", "sibling_school", "sibling_college"
        ],
        "optional_10th_marks": [
            "tenth_school_name", "tenth_year_of_passing", "tenth_board_of_study",
            "tenth_english", "tenth_tamil", "tenth_hindi", "tenth_maths",
            "tenth_science", "tenth_social_science", "tenth_total_marks"
        ],
        "optional_12th_marks": [
            "twelfth_school_name", "twelfth_year_of_passing", "twelfth_board_of_study",
            "twelfth_english", "twelfth_tamil", "twelfth_physics", "twelfth_chemistry",
            "twelfth_maths", "twelfth_biology", "twelfth_computer_science", "twelfth_total_marks"
        ],
        "optional_entrance_exam": [
            "entrance_exam_1", "entrance_exam_1_percentile", "entrance_exam_1_mark",
            "entrance_exam_2", "entrance_exam_2_percentile", "entrance_exam_2_mark",
            "entrance_exam_3", "entrance_exam_3_percentile", "entrance_exam_3_mark"
        ],
        "optional_counselling": [
            "counselling_forum_1", "counselling_round_1", "all_india_rank_1", "community_rank_1", "counselling_college_1",
            "counselling_forum_2", "counselling_round_2", "all_india_rank_2", "community_rank_2", "counselling_college_2",
            "counselling_forum_3", "counselling_round_3", "all_india_rank_3", "community_rank_3", "counselling_college_3"
        ]
    }
    
    return {
        "message": "Excel template column specifications",
        "columns": columns,
        "notes": [
            "Only 'student_id' and 'student_name' are required",
            "All other columns are optional",
            "Date format for 'dob': YYYY-MM-DD or DD/MM/YYYY",
            "batch_id will be provided during upload, not in Excel file"
        ]
    }


if __name__ == '__main__':
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)
