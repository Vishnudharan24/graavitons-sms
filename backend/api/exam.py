from fastapi import FastAPI, HTTPException, status, Depends, Query, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, validator, Field
from typing import List, Optional
import psycopg2
from psycopg2 import sql
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from io import BytesIO
from config import CORS_ORIGINS, APP_TITLE
from api.middleware import get_current_user
from db_pool import get_db_connection

app = FastAPI(title=APP_TITLE)

MOCK_SUBJECT_CONFIG = {
    "maths": {"aliases": {"maths", "mathematics", "math"}, "unit_field": "mathsUnitNames"},
    "physics": {"aliases": {"physics"}, "unit_field": "physicsUnitNames"},
    "chemistry": {"aliases": {"chemistry"}, "unit_field": "chemistryUnitNames"},
    "biology": {"aliases": {"biology"}, "unit_field": "biologyUnitNames"},
}

SUBJECT_CANONICAL = {
    "maths": "Mathematics",
    "math": "Mathematics",
    "mathematics": "Mathematics",
    "physics": "Physics",
    "chemistry": "Chemistry",
    "biology": "Biology",
}


def normalize_subject_label(value: str) -> str:
    key = str(value or "").strip().lower()
    if not key:
        return ""
    if key in SUBJECT_CANONICAL:
        return SUBJECT_CANONICAL[key]
    return " ".join(part.capitalize() for part in key.split())


def normalize_subject_key(value: str) -> str:
    label = normalize_subject_label(value)
    return label.strip().lower()


def get_batch_mock_subjects(batch_subjects):
    if not batch_subjects:
        return ["maths", "physics", "chemistry", "biology"]

    selected = []
    lowered = {str(s).strip().lower() for s in batch_subjects if str(s).strip()}
    for key in ["maths", "physics", "chemistry", "biology"]:
        aliases = MOCK_SUBJECT_CONFIG[key]["aliases"]
        if lowered.intersection(aliases):
            selected.append(key)

    return selected if selected else ["maths", "physics", "chemistry", "biology"]


def split_units(unit_text: str):
    if not unit_text:
        return []
    return [u.strip() for u in unit_text.split(',') if u.strip()]


def extract_grade_from_batch_name(batch_name: Optional[str]) -> Optional[int]:
    if not batch_name:
        return None
    import re
    grade_match = re.search(r'\d+', str(batch_name))
    if grade_match:
        return int(grade_match.group())
    return None


def normalized_subject_sql(column_name: str) -> str:
    return f"LOWER(TRIM({column_name}))"

# CORS configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# Pydantic models for Unit Test
class DailyTestStudentMark(BaseModel):
    id: str  # student_id
    marks: Optional[str] = None


class DailyTestCreate(BaseModel):
    batch_id: int
    examName: str
    examDate: date
    subject: str
    unitName: str
    totalMarks: int
    subjectTotalMarks: Optional[int] = None
    testTotalMarks: Optional[int] = None
    examType: str
    studentMarks: List[DailyTestStudentMark]


# Pydantic models for Monthly Test
class MockTestStudentMark(BaseModel):
    id: str  # student_id
    mathsMarks: Optional[str] = None
    physicsMarks: Optional[str] = None
    chemistryMarks: Optional[str] = None
    biologyMarks: Optional[str] = None


class MockTestCreate(BaseModel):
    batch_id: int
    examName: str
    examDate: date
    examType: str
    mathsUnitNames: str
    physicsUnitNames: str
    chemistryUnitNames: str
    biologyUnitNames: str
    mathsTotalMarks: Optional[int] = None
    physicsTotalMarks: Optional[int] = None
    chemistryTotalMarks: Optional[int] = None
    biologyTotalMarks: Optional[int] = None
    testTotalMarks: Optional[int] = None
    studentMarks: List[MockTestStudentMark]


class DailyTestBulkItem(BaseModel):
    examName: str
    examDate: date
    subject: str
    unitName: str
    totalMarks: int
    subjectTotalMarks: Optional[int] = None
    testTotalMarks: Optional[int] = None
    examType: Optional[str] = "daily test"
    studentMarks: List[DailyTestStudentMark]


class DailyTestBulkCreate(BaseModel):
    batch_id: int
    exams: List[DailyTestBulkItem]

    @validator('exams')
    def validate_daily_exams_not_empty(cls, value):
        if not value:
            raise ValueError("At least one unit test is required")
        return value


class MockTestBulkItem(BaseModel):
    examName: str
    examDate: date
    examType: Optional[str] = "mock test"
    mathsUnitNames: str = ""
    physicsUnitNames: str = ""
    chemistryUnitNames: str = ""
    biologyUnitNames: str = ""
    mathsTotalMarks: Optional[int] = None
    physicsTotalMarks: Optional[int] = None
    chemistryTotalMarks: Optional[int] = None
    biologyTotalMarks: Optional[int] = None
    testTotalMarks: Optional[int] = None
    studentMarks: List[MockTestStudentMark]


class MockTestBulkCreate(BaseModel):
    batch_id: int
    exams: List[MockTestBulkItem]

    @validator('exams')
    def validate_mock_exams_not_empty(cls, value):
        if not value:
            raise ValueError("At least one monthly test is required")
        return value


class DailyTestGroupRef(BaseModel):
    test_date: date
    subject: str
    unit_name: str
    subject_total_marks: Optional[int] = None
    test_total_marks: Optional[int] = None


class DailyTestMarkUpdate(BaseModel):
    student_id: str
    marks: Optional[str] = None


class DailyTestGroupUpdate(BaseModel):
    test_date: date
    subject: str
    unit_name: str
    subject_total_marks: Optional[int] = None
    test_total_marks: Optional[int] = None
    studentMarks: List[DailyTestMarkUpdate]


class MockTestGroupRef(BaseModel):
    test_date: date
    maths_unit_names: List[str] = Field(default_factory=list)
    physics_unit_names: List[str] = Field(default_factory=list)
    chemistry_unit_names: List[str] = Field(default_factory=list)
    biology_unit_names: List[str] = Field(default_factory=list)
    maths_total_marks: Optional[int] = None
    physics_total_marks: Optional[int] = None
    chemistry_total_marks: Optional[int] = None
    biology_total_marks: Optional[int] = None
    test_total_marks: Optional[int] = None


class MockTestMarkUpdate(BaseModel):
    student_id: str
    maths_marks: Optional[str] = None
    physics_marks: Optional[str] = None
    chemistry_marks: Optional[str] = None
    biology_marks: Optional[str] = None


class MockTestGroupUpdate(BaseModel):
    test_date: date
    maths_unit_names: List[str] = Field(default_factory=list)
    physics_unit_names: List[str] = Field(default_factory=list)
    chemistry_unit_names: List[str] = Field(default_factory=list)
    biology_unit_names: List[str] = Field(default_factory=list)
    maths_total_marks: Optional[int] = None
    physics_total_marks: Optional[int] = None
    chemistry_total_marks: Optional[int] = None
    biology_total_marks: Optional[int] = None
    test_total_marks: Optional[int] = None
    original_maths_total_marks: Optional[int] = None
    original_physics_total_marks: Optional[int] = None
    original_chemistry_total_marks: Optional[int] = None
    original_biology_total_marks: Optional[int] = None
    original_test_total_marks: Optional[int] = None
    studentMarks: List[MockTestMarkUpdate]


@app.post("/api/exam/daily-test", status_code=status.HTTP_201_CREATED)
async def create_daily_test(exam_data: DailyTestCreate, current_user: dict = Depends(get_current_user)):
    """
    Create unit test marks for students
    """
    conn = None
    cursor = None
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get batch details to extract grade and branch
        cursor.execute("""
            SELECT batch_name, type, subjects FROM batch WHERE batch_id = %s
        """, (exam_data.batch_id,))
        
        batch_result = cursor.fetchone()
        if not batch_result:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Batch with ID {exam_data.batch_id} not found"
            )
        
        batch_name, batch_type, batch_subjects = batch_result
        normalized_subject = normalize_subject_label(exam_data.subject)
        if not normalized_subject:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Subject is required"
            )

        # If batch has configured subjects, enforce subject membership (case-insensitive, alias-safe)
        normalized_batch_subject_keys = {
            normalize_subject_key(s) for s in (batch_subjects or []) if str(s).strip()
        }
        if normalized_batch_subject_keys and normalize_subject_key(normalized_subject) not in normalized_batch_subject_keys:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Subject '{exam_data.subject}' is not configured for this batch"
            )
        
        # Extract grade from batch (assuming format like "Grade 11", "12th Grade", etc.)
        grade = None
        if batch_name:
            # Try to extract numeric grade
            import re
            grade_match = re.search(r'\d+', batch_name)
            if grade_match:
                grade = int(grade_match.group())
        
        # Get branch from student (we'll fetch it for each student)
        inserted_count = 0
        failed_students = []
        subject_total_marks = exam_data.subjectTotalMarks if exam_data.subjectTotalMarks is not None else exam_data.totalMarks
        test_total_marks = exam_data.testTotalMarks if exam_data.testTotalMarks is not None else subject_total_marks
        
        for student_mark in exam_data.studentMarks:
            try:
                # Get student's branch
                cursor.execute("""
                    SELECT student_no, branch
                    FROM student
                    WHERE student_id = %s AND batch_id = %s
                    ORDER BY created_at DESC, student_no DESC
                    LIMIT 1
                """, (student_mark.id, exam_data.batch_id))
                
                student_result = cursor.fetchone()
                if not student_result:
                    failed_students.append({
                        "student_id": student_mark.id,
                        "reason": "Student not found"
                    })
                    continue
                
                student_no = student_result[0]
                branch = student_result[1]
                
                # Store marks as-is (supports integers, 'A' for absent, '-' for N/A, negative marks)
                marks = student_mark.marks.strip() if student_mark.marks and student_mark.marks.strip() else None
                
                # Insert unit test record
                cursor.execute("""
                    INSERT INTO daily_test (
                        student_no, grade, branch, test_date, 
                        subject, unit_name, total_marks, subject_total_marks, test_total_marks
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    student_no,
                    grade,
                    branch,
                    exam_data.examDate,
                    normalized_subject,
                    exam_data.unitName,
                    marks,
                    subject_total_marks,
                    test_total_marks
                ))
                
                inserted_count += 1
                
            except psycopg2.Error as db_error:
                failed_students.append({
                    "student_id": student_mark.id,
                    "reason": f"Database error: {str(db_error)}"
                })
        
        conn.commit()
        
        response = {
            "message": "Unit test marks added successfully",
            "exam_name": exam_data.examName,
            "exam_date": str(exam_data.examDate),
            "subject": normalized_subject,
            "unit_name": exam_data.unitName,
            "total_marks": exam_data.totalMarks,
            "subject_total_marks": subject_total_marks,
            "test_total_marks": test_total_marks,
            "inserted_count": inserted_count,
            "total_students": len(exam_data.studentMarks)
        }
        
        if failed_students:
            response["failed_students"] = failed_students
        
        return response
        
    except HTTPException:
        raise
    except Exception as e:
        if conn:
            conn.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to create unit test: {str(e)}"
        )
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@app.post("/api/exam/mock-test", status_code=status.HTTP_201_CREATED)
async def create_mock_test(exam_data: MockTestCreate, current_user: dict = Depends(get_current_user)):
    """
    Create monthly test marks for students
    """
    conn = None
    cursor = None
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get batch details to extract grade and configured subjects
        cursor.execute("""
            SELECT batch_name, type, subjects FROM batch WHERE batch_id = %s
        """, (exam_data.batch_id,))
        
        batch_result = cursor.fetchone()
        if not batch_result:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Batch with ID {exam_data.batch_id} not found"
            )
        
        batch_name, batch_type, batch_subjects = batch_result
        active_subjects = set(get_batch_mock_subjects(batch_subjects))
        
        # Extract grade from batch
        grade = None
        if batch_name:
            import re
            grade_match = re.search(r'\d+', batch_name)
            if grade_match:
                grade = int(grade_match.group())
        
        # Parse unit names only for subjects configured in the batch
        maths_units = split_units(exam_data.mathsUnitNames) if "maths" in active_subjects else []
        physics_units = split_units(exam_data.physicsUnitNames) if "physics" in active_subjects else []
        chemistry_units = split_units(exam_data.chemistryUnitNames) if "chemistry" in active_subjects else []
        biology_units = split_units(exam_data.biologyUnitNames) if "biology" in active_subjects else []

        maths_total_marks = exam_data.mathsTotalMarks if "maths" in active_subjects else None
        physics_total_marks = exam_data.physicsTotalMarks if "physics" in active_subjects else None
        chemistry_total_marks = exam_data.chemistryTotalMarks if "chemistry" in active_subjects else None
        biology_total_marks = exam_data.biologyTotalMarks if "biology" in active_subjects else None
        test_total_marks = exam_data.testTotalMarks
        if test_total_marks is None:
            total_parts = [v for v in [maths_total_marks, physics_total_marks, chemistry_total_marks, biology_total_marks] if isinstance(v, int)]
            test_total_marks = sum(total_parts) if total_parts else None
        
        inserted_count = 0
        failed_students = []
        
        for student_mark in exam_data.studentMarks:
            try:
                # Get student's branch
                cursor.execute("""
                    SELECT student_no, branch
                    FROM student
                    WHERE student_id = %s AND batch_id = %s
                    ORDER BY created_at DESC, student_no DESC
                    LIMIT 1
                """, (student_mark.id, exam_data.batch_id))
                
                student_result = cursor.fetchone()
                if not student_result:
                    failed_students.append({
                        "student_id": student_mark.id,
                        "reason": "Student not found"
                    })
                    continue
                
                student_no = student_result[0]
                branch = student_result[1]
                
                # Store marks as-is (supports integers, 'A' for absent, '-' for N/A, negative marks)
                maths_marks = student_mark.mathsMarks.strip() if "maths" in active_subjects and student_mark.mathsMarks and student_mark.mathsMarks.strip() else None
                physics_marks = student_mark.physicsMarks.strip() if "physics" in active_subjects and student_mark.physicsMarks and student_mark.physicsMarks.strip() else None
                chemistry_marks = student_mark.chemistryMarks.strip() if "chemistry" in active_subjects and student_mark.chemistryMarks and student_mark.chemistryMarks.strip() else None
                biology_marks = student_mark.biologyMarks.strip() if "biology" in active_subjects and student_mark.biologyMarks and student_mark.biologyMarks.strip() else None
                
                # Calculate total marks only from numeric values
                def safe_int(val):
                    try:
                        return int(val)
                    except (ValueError, TypeError):
                        return None
                
                numeric_marks = [safe_int(m) for m in [maths_marks, physics_marks, chemistry_marks, biology_marks]]
                valid_marks = [m for m in numeric_marks if m is not None]
                total_marks = str(sum(valid_marks)) if valid_marks else None
                
                # Insert monthly test record
                cursor.execute("""
                    INSERT INTO mock_test (
                        student_no, grade, branch, test_date,
                        maths_marks, physics_marks, chemistry_marks, biology_marks,
                        maths_unit_names, physics_unit_names, chemistry_unit_names, biology_unit_names,
                        total_marks,
                        maths_total_marks, physics_total_marks, chemistry_total_marks, biology_total_marks,
                        test_total_marks
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    student_no,
                    grade,
                    branch,
                    exam_data.examDate,
                    maths_marks,
                    physics_marks,
                    chemistry_marks,
                    biology_marks,
                    maths_units,
                    physics_units,
                    chemistry_units,
                    biology_units,
                    total_marks,
                    maths_total_marks,
                    physics_total_marks,
                    chemistry_total_marks,
                    biology_total_marks,
                    test_total_marks
                ))
                
                inserted_count += 1
                
            except ValueError as ve:
                failed_students.append({
                    "student_id": student_mark.id,
                    "reason": f"Invalid marks value"
                })
            except psycopg2.Error as db_error:
                failed_students.append({
                    "student_id": student_mark.id,
                    "reason": f"Database error: {str(db_error)}"
                })
        
        conn.commit()
        
        response = {
            "message": "Monthly test marks added successfully",
            "exam_name": exam_data.examName,
            "exam_date": str(exam_data.examDate),
            "active_subjects": list(active_subjects),
            "units": {
                "maths": maths_units,
                "physics": physics_units,
                "chemistry": chemistry_units,
                "biology": biology_units
            },
            "subject_total_marks": {
                "maths": maths_total_marks,
                "physics": physics_total_marks,
                "chemistry": chemistry_total_marks,
                "biology": biology_total_marks
            },
            "test_total_marks": test_total_marks,
            "inserted_count": inserted_count,
            "total_students": len(exam_data.studentMarks)
        }
        
        if failed_students:
            response["failed_students"] = failed_students
        
        return response
        
    except HTTPException:
        raise
    except Exception as e:
        if conn:
            conn.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to create monthly test: {str(e)}"
        )
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@app.post("/api/exam/daily-test/bulk", status_code=status.HTTP_201_CREATED)
async def create_daily_test_bulk(exam_data: DailyTestBulkCreate, current_user: dict = Depends(get_current_user)):
    """
    Upload multiple unit tests in a single request for one batch.
    """
    results = []
    failed_exams = []
    total_inserted = 0

    for index, exam in enumerate(exam_data.exams, start=1):
        single_payload = DailyTestCreate(
            batch_id=exam_data.batch_id,
            examName=exam.examName,
            examDate=exam.examDate,
            subject=exam.subject,
            unitName=exam.unitName,
            totalMarks=exam.totalMarks,
            subjectTotalMarks=exam.subjectTotalMarks,
            testTotalMarks=exam.testTotalMarks,
            examType=exam.examType or "daily test",
            studentMarks=exam.studentMarks
        )

        try:
            result = await create_daily_test(single_payload, current_user)
            inserted_count = result.get("inserted_count", 0)
            total_inserted += inserted_count
            results.append({
                "index": index,
                "exam_name": exam.examName,
                "exam_date": str(exam.examDate),
                "status": "success",
                "inserted_count": inserted_count,
                "total_students": result.get("total_students", len(exam.studentMarks)),
                "failed_students": result.get("failed_students", [])
            })
        except HTTPException as e:
            failed_exams.append({
                "index": index,
                "exam_name": exam.examName,
                "exam_date": str(exam.examDate),
                "status": "failed",
                "reason": e.detail
            })

    return {
        "message": "Bulk unit test upload completed",
        "batch_id": exam_data.batch_id,
        "total_exams": len(exam_data.exams),
        "successful_exams": len(results),
        "failed_exams_count": len(failed_exams),
        "total_inserted_records": total_inserted,
        "results": results,
        "failed_exams": failed_exams
    }


@app.post("/api/exam/mock-test/bulk", status_code=status.HTTP_201_CREATED)
async def create_mock_test_bulk(exam_data: MockTestBulkCreate, current_user: dict = Depends(get_current_user)):
    """
    Upload multiple monthly tests in a single request for one batch.
    """
    results = []
    failed_exams = []
    total_inserted = 0

    for index, exam in enumerate(exam_data.exams, start=1):
        single_payload = MockTestCreate(
            batch_id=exam_data.batch_id,
            examName=exam.examName,
            examDate=exam.examDate,
            examType=exam.examType or "mock test",
            mathsUnitNames=exam.mathsUnitNames,
            physicsUnitNames=exam.physicsUnitNames,
            chemistryUnitNames=exam.chemistryUnitNames,
            biologyUnitNames=exam.biologyUnitNames,
            mathsTotalMarks=exam.mathsTotalMarks,
            physicsTotalMarks=exam.physicsTotalMarks,
            chemistryTotalMarks=exam.chemistryTotalMarks,
            biologyTotalMarks=exam.biologyTotalMarks,
            testTotalMarks=exam.testTotalMarks,
            studentMarks=exam.studentMarks
        )

        try:
            result = await create_mock_test(single_payload, current_user)
            inserted_count = result.get("inserted_count", 0)
            total_inserted += inserted_count
            results.append({
                "index": index,
                "exam_name": exam.examName,
                "exam_date": str(exam.examDate),
                "status": "success",
                "inserted_count": inserted_count,
                "total_students": result.get("total_students", len(exam.studentMarks)),
                "failed_students": result.get("failed_students", [])
            })
        except HTTPException as e:
            failed_exams.append({
                "index": index,
                "exam_name": exam.examName,
                "exam_date": str(exam.examDate),
                "status": "failed",
                "reason": e.detail
            })

    return {
        "message": "Bulk monthly test upload completed",
        "batch_id": exam_data.batch_id,
        "total_exams": len(exam_data.exams),
        "successful_exams": len(results),
        "failed_exams_count": len(failed_exams),
        "total_inserted_records": total_inserted,
        "results": results,
        "failed_exams": failed_exams
    }


@app.get("/api/exam/daily-test/batch/{batch_id}/groups")
async def get_daily_test_groups(batch_id: int, current_user: dict = Depends(get_current_user)):
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT batch_id FROM batch WHERE batch_id = %s", (batch_id,))
        if not cursor.fetchone():
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Batch with ID {batch_id} not found")

        cursor.execute("""
            SELECT
                dt.test_date,
                dt.subject,
                dt.unit_name,
                dt.subject_total_marks,
                dt.test_total_marks,
                COUNT(*) AS entries_count,
                COUNT(DISTINCT dt.student_no) AS student_count,
                MIN(dt.created_at) AS created_at
            FROM daily_test dt
            JOIN student s ON s.student_no = dt.student_no
            WHERE s.batch_id = %s
            GROUP BY dt.test_date, dt.subject, dt.unit_name, dt.subject_total_marks, dt.test_total_marks
            ORDER BY dt.test_date DESC, dt.subject, dt.unit_name
        """, (batch_id,))

        groups = []
        for row in cursor.fetchall():
            groups.append({
                "test_date": row[0].isoformat() if row[0] else None,
                "subject": row[1],
                "unit_name": row[2],
                "subject_total_marks": row[3],
                "test_total_marks": row[4],
                "entries_count": row[5],
                "student_count": row[6],
                "created_at": row[7].isoformat() if row[7] else None,
            })

        return {"groups": groups, "total_groups": len(groups)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch unit test groups: {str(e)}"
        )
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@app.post("/api/exam/daily-test/batch/{batch_id}/records")
async def get_daily_test_group_records(
    batch_id: int,
    group_ref: DailyTestGroupRef,
    current_user: dict = Depends(get_current_user)
):
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        normalized_subject = normalize_subject_key(group_ref.subject)
        cursor.execute(f"""
            SELECT
                s.student_id,
                s.student_name,
                COALESCE(dt.total_marks, '') AS marks
            FROM student s
            LEFT JOIN daily_test dt
                ON dt.student_no = s.student_no
                AND dt.test_date = %s
                AND COALESCE(dt.unit_name, '') = COALESCE(%s, '')
                AND {normalized_subject_sql('dt.subject')} = %s
            WHERE s.batch_id = %s
            ORDER BY
                CASE WHEN s.student_id ~ '^[0-9]+$' THEN 0 ELSE 1 END,
                CASE WHEN s.student_id ~ '^[0-9]+$' THEN s.student_id::BIGINT END,
                s.student_id ASC,
                s.student_name ASC
        """, (group_ref.test_date, group_ref.unit_name, normalized_subject, batch_id))

        rows = cursor.fetchall()
        return {
            "records": [
                {
                    "student_id": r[0],
                    "student_name": r[1],
                    "marks": r[2]
                }
                for r in rows
            ],
            "total_records": len(rows)
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch unit test records: {str(e)}"
        )
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@app.put("/api/exam/daily-test/batch/{batch_id}")
async def update_daily_test_group(
    batch_id: int,
    payload: DailyTestGroupUpdate,
    current_user: dict = Depends(get_current_user)
):
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT batch_name FROM batch WHERE batch_id = %s", (batch_id,))
        batch_row = cursor.fetchone()
        if not batch_row:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Batch with ID {batch_id} not found")

        batch_grade = extract_grade_from_batch_name(batch_row[0])
        normalized_subject_label = normalize_subject_label(payload.subject)
        normalized_subject = normalize_subject_key(payload.subject)

        cursor.execute("""
            SELECT student_id, student_no, branch, grade
            FROM student
            WHERE batch_id = %s
        """, (batch_id,))
        student_map = {r[0]: {"student_no": r[1], "branch": r[2], "grade": r[3]} for r in cursor.fetchall()}

        updated_count = 0
        inserted_count = 0
        deleted_count = 0
        skipped_count = 0

        for student_mark in payload.studentMarks:
            sid = student_mark.student_id
            if sid not in student_map:
                skipped_count += 1
                continue

            marks = (student_mark.marks or '').strip()
            marks_value = marks if marks else None

            cursor.execute(f"""
                SELECT test_id
                FROM daily_test
                                WHERE student_no = %s
                  AND test_date = %s
                  AND COALESCE(unit_name, '') = COALESCE(%s, '')
                  AND {normalized_subject_sql('subject')} = %s
                LIMIT 1
                        """, (student_map[sid]["student_no"], payload.test_date, payload.unit_name, normalized_subject))
            existing = cursor.fetchone()

            if existing:
                cursor.execute("""
                    UPDATE daily_test
                    SET
                        total_marks = %s,
                        subject_total_marks = %s,
                        test_total_marks = %s
                    WHERE test_id = %s
                """, (marks_value, payload.subject_total_marks, payload.test_total_marks, existing[0]))
                updated_count += 1
            else:
                student_meta = student_map[sid]
                cursor.execute("""
                    INSERT INTO daily_test (
                        student_no, grade, branch, test_date,
                        subject, unit_name, total_marks, subject_total_marks, test_total_marks
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    student_meta["student_no"],
                    student_meta["grade"] if student_meta["grade"] is not None else batch_grade,
                    student_meta["branch"],
                    payload.test_date,
                    normalized_subject_label,
                    payload.unit_name,
                    marks_value,
                    payload.subject_total_marks,
                    payload.test_total_marks
                ))
                inserted_count += 1

        conn.commit()

        return {
            "message": "Unit test marks updated successfully",
            "updated_count": updated_count,
            "inserted_count": inserted_count,
            "deleted_count": deleted_count,
            "skipped_count": skipped_count
        }
    except HTTPException:
        raise
    except Exception as e:
        if conn:
            conn.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update unit test marks: {str(e)}"
        )
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@app.delete("/api/exam/daily-test/batch/{batch_id}")
async def delete_daily_test_group(
    batch_id: int,
    payload: DailyTestGroupRef,
    current_user: dict = Depends(get_current_user)
):
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        normalized_subject = normalize_subject_key(payload.subject)
        cursor.execute(f"""
            DELETE FROM daily_test dt
            USING student s
                        WHERE dt.student_no = s.student_no
              AND s.batch_id = %s
              AND dt.test_date = %s
              AND COALESCE(dt.unit_name, '') = COALESCE(%s, '')
              AND {normalized_subject_sql('dt.subject')} = %s
        """, (batch_id, payload.test_date, payload.unit_name, normalized_subject))
        deleted_count = cursor.rowcount

        conn.commit()
        return {
            "message": "Unit test deleted successfully",
            "deleted_count": deleted_count
        }
    except Exception as e:
        if conn:
            conn.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete unit test group: {str(e)}"
        )
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@app.get("/api/exam/mock-test/batch/{batch_id}/groups")
async def get_mock_test_groups(batch_id: int, current_user: dict = Depends(get_current_user)):
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT subjects FROM batch WHERE batch_id = %s", (batch_id,))
        batch_row = cursor.fetchone()
        if not batch_row:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Batch with ID {batch_id} not found")

        active_subjects = get_batch_mock_subjects(batch_row[0])

        cursor.execute("""
            SELECT
                mt.test_date,
                COALESCE(mt.maths_unit_names, ARRAY[]::text[]) AS maths_unit_names,
                COALESCE(mt.physics_unit_names, ARRAY[]::text[]) AS physics_unit_names,
                COALESCE(mt.chemistry_unit_names, ARRAY[]::text[]) AS chemistry_unit_names,
                COALESCE(mt.biology_unit_names, ARRAY[]::text[]) AS biology_unit_names,
                mt.maths_total_marks,
                mt.physics_total_marks,
                mt.chemistry_total_marks,
                mt.biology_total_marks,
                mt.test_total_marks,
                COUNT(*) AS entries_count,
                COUNT(DISTINCT mt.student_no) AS student_count,
                MIN(mt.created_at) AS created_at
            FROM mock_test mt
            JOIN student s ON s.student_no = mt.student_no
            WHERE s.batch_id = %s
            GROUP BY
                mt.test_date,
                mt.maths_unit_names,
                mt.physics_unit_names,
                mt.chemistry_unit_names,
                mt.biology_unit_names,
                mt.maths_total_marks,
                mt.physics_total_marks,
                mt.chemistry_total_marks,
                mt.biology_total_marks,
                mt.test_total_marks
            ORDER BY mt.test_date DESC
        """, (batch_id,))

        groups = []
        for row in cursor.fetchall():
            groups.append({
                "test_date": row[0].isoformat() if row[0] else None,
                "maths_unit_names": row[1] or [],
                "physics_unit_names": row[2] or [],
                "chemistry_unit_names": row[3] or [],
                "biology_unit_names": row[4] or [],
                "maths_total_marks": row[5],
                "physics_total_marks": row[6],
                "chemistry_total_marks": row[7],
                "biology_total_marks": row[8],
                "test_total_marks": row[9],
                "entries_count": row[10],
                "student_count": row[11],
                "created_at": row[12].isoformat() if row[12] else None,
            })

        return {"groups": groups, "active_subjects": active_subjects, "total_groups": len(groups)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch monthly test groups: {str(e)}"
        )
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@app.post("/api/exam/mock-test/batch/{batch_id}/records")
async def get_mock_test_group_records(
    batch_id: int,
    group_ref: MockTestGroupRef,
    current_user: dict = Depends(get_current_user)
):
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT
                s.student_id,
                s.student_name,
                COALESCE(mt.maths_marks, '') AS maths_marks,
                COALESCE(mt.physics_marks, '') AS physics_marks,
                COALESCE(mt.chemistry_marks, '') AS chemistry_marks,
                COALESCE(mt.biology_marks, '') AS biology_marks
            FROM student s
            LEFT JOIN mock_test mt
                ON mt.student_no = s.student_no
                AND mt.test_date = %s
                AND COALESCE(mt.maths_unit_names, ARRAY[]::text[]) = %s
                AND COALESCE(mt.physics_unit_names, ARRAY[]::text[]) = %s
                AND COALESCE(mt.chemistry_unit_names, ARRAY[]::text[]) = %s
                AND COALESCE(mt.biology_unit_names, ARRAY[]::text[]) = %s
                AND mt.maths_total_marks IS NOT DISTINCT FROM %s
                AND mt.physics_total_marks IS NOT DISTINCT FROM %s
                AND mt.chemistry_total_marks IS NOT DISTINCT FROM %s
                AND mt.biology_total_marks IS NOT DISTINCT FROM %s
                AND mt.test_total_marks IS NOT DISTINCT FROM %s
            WHERE s.batch_id = %s
            ORDER BY
                CASE WHEN s.student_id ~ '^[0-9]+$' THEN 0 ELSE 1 END,
                CASE WHEN s.student_id ~ '^[0-9]+$' THEN s.student_id::BIGINT END,
                s.student_id ASC,
                s.student_name ASC
        """, (
            group_ref.test_date,
            group_ref.maths_unit_names,
            group_ref.physics_unit_names,
            group_ref.chemistry_unit_names,
            group_ref.biology_unit_names,
            group_ref.maths_total_marks,
            group_ref.physics_total_marks,
            group_ref.chemistry_total_marks,
            group_ref.biology_total_marks,
            group_ref.test_total_marks,
            batch_id
        ))

        rows = cursor.fetchall()
        return {
            "records": [
                {
                    "student_id": r[0],
                    "student_name": r[1],
                    "maths_marks": r[2],
                    "physics_marks": r[3],
                    "chemistry_marks": r[4],
                    "biology_marks": r[5],
                }
                for r in rows
            ],
            "total_records": len(rows)
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch monthly test records: {str(e)}"
        )
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@app.put("/api/exam/mock-test/batch/{batch_id}")
async def update_mock_test_group(
    batch_id: int,
    payload: MockTestGroupUpdate,
    current_user: dict = Depends(get_current_user)
):
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT batch_name, subjects FROM batch WHERE batch_id = %s", (batch_id,))
        batch_row = cursor.fetchone()
        if not batch_row:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Batch with ID {batch_id} not found")

        batch_grade = extract_grade_from_batch_name(batch_row[0])
        active_subjects = set(get_batch_mock_subjects(batch_row[1]))

        cursor.execute("""
            SELECT student_id, student_no, branch, grade
            FROM student
            WHERE batch_id = %s
        """, (batch_id,))
        student_map = {r[0]: {"student_no": r[1], "branch": r[2], "grade": r[3]} for r in cursor.fetchall()}

        def safe_int(val):
            try:
                return int(val)
            except (TypeError, ValueError):
                return None

        updated_count = 0
        inserted_count = 0
        deleted_count = 0
        skipped_count = 0

        match_maths_total = payload.original_maths_total_marks if payload.original_maths_total_marks is not None else payload.maths_total_marks
        match_physics_total = payload.original_physics_total_marks if payload.original_physics_total_marks is not None else payload.physics_total_marks
        match_chemistry_total = payload.original_chemistry_total_marks if payload.original_chemistry_total_marks is not None else payload.chemistry_total_marks
        match_biology_total = payload.original_biology_total_marks if payload.original_biology_total_marks is not None else payload.biology_total_marks
        match_test_total = payload.original_test_total_marks if payload.original_test_total_marks is not None else payload.test_total_marks

        for student_mark in payload.studentMarks:
            sid = student_mark.student_id
            if sid not in student_map:
                skipped_count += 1
                continue

            maths_raw = (student_mark.maths_marks or '').strip() if 'maths' in active_subjects else ''
            physics_raw = (student_mark.physics_marks or '').strip() if 'physics' in active_subjects else ''
            chemistry_raw = (student_mark.chemistry_marks or '').strip() if 'chemistry' in active_subjects else ''
            biology_raw = (student_mark.biology_marks or '').strip() if 'biology' in active_subjects else ''

            maths_marks = maths_raw if maths_raw else None
            physics_marks = physics_raw if physics_raw else None
            chemistry_marks = chemistry_raw if chemistry_raw else None
            biology_marks = biology_raw if biology_raw else None

            cursor.execute("""
                SELECT test_id
                FROM mock_test
                                WHERE student_no = %s
                  AND test_date = %s
                  AND COALESCE(maths_unit_names, ARRAY[]::text[]) = %s
                  AND COALESCE(physics_unit_names, ARRAY[]::text[]) = %s
                  AND COALESCE(chemistry_unit_names, ARRAY[]::text[]) = %s
                  AND COALESCE(biology_unit_names, ARRAY[]::text[]) = %s
                  AND maths_total_marks IS NOT DISTINCT FROM %s
                  AND physics_total_marks IS NOT DISTINCT FROM %s
                  AND chemistry_total_marks IS NOT DISTINCT FROM %s
                  AND biology_total_marks IS NOT DISTINCT FROM %s
                  AND test_total_marks IS NOT DISTINCT FROM %s
                LIMIT 1
            """, (
                student_map[sid]["student_no"],
                payload.test_date,
                payload.maths_unit_names,
                payload.physics_unit_names,
                payload.chemistry_unit_names,
                payload.biology_unit_names,
                                match_maths_total,
                                match_physics_total,
                                match_chemistry_total,
                                match_biology_total,
                                match_test_total,
            ))
            existing = cursor.fetchone()

            numeric_marks = [safe_int(v) for v in [maths_marks, physics_marks, chemistry_marks, biology_marks]]
            valid_marks = [m for m in numeric_marks if m is not None]
            total_marks = str(sum(valid_marks)) if valid_marks else None

            if existing:
                cursor.execute("""
                    UPDATE mock_test
                    SET
                        maths_marks = %s,
                        physics_marks = %s,
                        chemistry_marks = %s,
                        biology_marks = %s,
                        total_marks = %s,
                        maths_total_marks = %s,
                        physics_total_marks = %s,
                        chemistry_total_marks = %s,
                        biology_total_marks = %s,
                        test_total_marks = %s
                    WHERE test_id = %s
                """, (
                    maths_marks,
                    physics_marks,
                    chemistry_marks,
                    biology_marks,
                    total_marks,
                    payload.maths_total_marks,
                    payload.physics_total_marks,
                    payload.chemistry_total_marks,
                    payload.biology_total_marks,
                    payload.test_total_marks,
                    existing[0]
                ))
                updated_count += 1
            else:
                student_meta = student_map[sid]
                cursor.execute("""
                    INSERT INTO mock_test (
                        student_no, grade, branch, test_date,
                        maths_marks, physics_marks, chemistry_marks, biology_marks,
                        maths_unit_names, physics_unit_names, chemistry_unit_names, biology_unit_names,
                        total_marks,
                        maths_total_marks, physics_total_marks, chemistry_total_marks, biology_total_marks,
                        test_total_marks
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    student_meta["student_no"],
                    student_meta["grade"] if student_meta["grade"] is not None else batch_grade,
                    student_meta["branch"],
                    payload.test_date,
                    maths_marks,
                    physics_marks,
                    chemistry_marks,
                    biology_marks,
                    payload.maths_unit_names,
                    payload.physics_unit_names,
                    payload.chemistry_unit_names,
                    payload.biology_unit_names,
                    total_marks,
                    payload.maths_total_marks,
                    payload.physics_total_marks,
                    payload.chemistry_total_marks,
                    payload.biology_total_marks,
                    payload.test_total_marks,
                ))
                inserted_count += 1

        conn.commit()
        return {
            "message": "Monthly test marks updated successfully",
            "updated_count": updated_count,
            "inserted_count": inserted_count,
            "deleted_count": deleted_count,
            "skipped_count": skipped_count
        }
    except HTTPException:
        raise
    except Exception as e:
        if conn:
            conn.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update monthly test marks: {str(e)}"
        )
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@app.delete("/api/exam/mock-test/batch/{batch_id}")
async def delete_mock_test_group(
    batch_id: int,
    payload: MockTestGroupRef,
    current_user: dict = Depends(get_current_user)
):
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            DELETE FROM mock_test mt
            USING student s
                        WHERE mt.student_no = s.student_no
              AND s.batch_id = %s
              AND mt.test_date = %s
              AND COALESCE(mt.maths_unit_names, ARRAY[]::text[]) = %s
              AND COALESCE(mt.physics_unit_names, ARRAY[]::text[]) = %s
              AND COALESCE(mt.chemistry_unit_names, ARRAY[]::text[]) = %s
              AND COALESCE(mt.biology_unit_names, ARRAY[]::text[]) = %s
              AND mt.maths_total_marks IS NOT DISTINCT FROM %s
              AND mt.physics_total_marks IS NOT DISTINCT FROM %s
              AND mt.chemistry_total_marks IS NOT DISTINCT FROM %s
              AND mt.biology_total_marks IS NOT DISTINCT FROM %s
              AND mt.test_total_marks IS NOT DISTINCT FROM %s
        """, (
            batch_id,
            payload.test_date,
            payload.maths_unit_names,
            payload.physics_unit_names,
            payload.chemistry_unit_names,
            payload.biology_unit_names,
            payload.maths_total_marks,
            payload.physics_total_marks,
            payload.chemistry_total_marks,
            payload.biology_total_marks,
            payload.test_total_marks,
        ))
        deleted_count = cursor.rowcount

        conn.commit()
        return {
            "message": "Monthly test deleted successfully",
            "deleted_count": deleted_count
        }
    except Exception as e:
        if conn:
            conn.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete monthly test group: {str(e)}"
        )
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@app.get("/api/exam/health")
async def health_check(current_user: dict = Depends(get_current_user)):
    """Health check endpoint"""
    return {"status": "healthy", "service": "exam-api"}


# ── Shared style helpers ─────────────────────────────────────────────────────

def _daily_styles():
    thin  = Side(style='thin',   color="BFBFBF")
    thick = Side(style='medium', color="1F3864")
    return {
        "col_header_fill":  PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid"),
        "meta_label_fill":  PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
        "meta_edit_fill":   PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "student_fill":     PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        "alt_student_fill": PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
        "marks_fill":       PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "col_header_font":  Font(bold=True, color="FFFFFF",  size=11, name="Calibri"),
        "meta_label_font":  Font(bold=True, color="1F3864",  size=10, name="Calibri"),
        "meta_hint_font":   Font(color="9E9E9E", size=9, name="Calibri", italic=True),
        "student_font":     Font(size=10, name="Calibri"),
        "id_font":          Font(size=10, name="Calibri", color="595959"),
        "border":           Border(left=thin, right=thin, top=thin, bottom=thin),
        "top_border":       Border(left=thin, right=thin, top=thick, bottom=thin),
        "bot_border":       Border(left=thin, right=thin, top=thin,  bottom=thick),
    }


# ─────────────────────────────────────────────────────────────────────────────
# 1. TEMPLATE GENERATOR  (replaces existing get_daily_test_template)
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/exam/template/daily-test/{batch_id}")
async def get_daily_test_template(
    batch_id: int,
    total_marks: int = 100,
    multi_template: bool = Query(False),
    test_count: int = Query(1, ge=1, le=100),
    current_user: dict = Depends(get_current_user)
):
    """
    Generate and download Excel template for unit test marks entry.

    multi_template=false  →  simple 3-col sheet (Admission No, Name, Marks)
    multi_template=true   →  clean multi-test sheet with coloured META rows
                              (fill Date/Subject/Unit once per test block,
                               then just Marks per student)
    """
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT student_id, student_name
            FROM student
            WHERE batch_id = %s
            ORDER BY
                CASE WHEN student_id ~ '^[0-9]+$' THEN 0 ELSE 1 END,
                CASE WHEN student_id ~ '^[0-9]+$' THEN student_id::BIGINT END,
                student_id ASC, student_name ASC
        """, (batch_id,))
        students = cursor.fetchall()

        if not students:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"No students found for batch ID {batch_id}"
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Unit Test Marks"

        if not multi_template:
            # ── Simple single-test template (unchanged) ──────────────────────
            hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            hfont = Font(bold=True, color="FFFFFF", size=12)
            thin  = Side(style='thin')
            bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)

            for c, h in enumerate(['Admission Number', 'Student Name', 'Marks'], 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.fill = hfill; cell.font = hfont
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = bdr

            for row, (sid, sname) in enumerate(students, 2):
                ws.cell(row=row, column=1, value=sid).border = bdr
                ws.cell(row=row, column=2, value=sname).border = bdr
                ws.cell(row=row, column=3, value="").border = bdr

            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 16

        else:
            # ── Clean multi-test template ────────────────────────────────────
            s = _daily_styles()

            COL_WIDTHS   = [10, 20, 32, 22, 20, 32, 14]
            COL_HEADERS  = [
                "Test No",
                "Admission No",
                "Student Name",
                "Exam Date\n(YYYY-MM-DD)",
                "Subject",
                "Topic / Unit Name",
                "Marks",
            ]
            TOTAL_COLS = len(COL_HEADERS)

            # Row 1: column headers
            ws.row_dimensions[1].height = 32
            for c, (header, width) in enumerate(zip(COL_HEADERS, COL_WIDTHS), 1):
                cell = ws.cell(row=1, column=c, value=header)
                cell.fill      = s["col_header_fill"]
                cell.font      = s["col_header_font"]
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border    = s["border"]
                ws.column_dimensions[get_column_letter(c)].width = width

            ws.freeze_panes = "A2"

            current_row = 2
            for test_no in range(1, test_count + 1):

                # ── META row ─────────────────────────────────────────────────
                ws.row_dimensions[current_row].height = 24

                # Col A: "TEST N" label
                c = ws.cell(row=current_row, column=1, value=f"TEST {test_no}")
                c.fill      = s["meta_label_fill"]
                c.font      = s["meta_label_font"]
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border    = s["top_border"]

                # Col B, C: empty label cells
                for col in (2, 3):
                    c = ws.cell(row=current_row, column=col, value="")
                    c.fill   = s["meta_label_fill"]
                    c.border = s["top_border"]

                # Cols D, E, F: yellow editable cells with placeholder hints
                hints = {4: "e.g. 2026-03-02", 5: "e.g. Maths", 6: "e.g. Continuity & Diff."}
                for col, hint in hints.items():
                    c = ws.cell(row=current_row, column=col, value="")
                    c.fill      = s["meta_edit_fill"]
                    c.font      = Font(color="BFBFBF", size=9, name="Calibri", italic=True)
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.border    = s["top_border"]
                    # Write hint as comment-style placeholder via number_format trick
                    # (actual hint shown via the value — user overwrites it)
                    c.value = hint

                # Col G: hint
                c = ws.cell(row=current_row, column=7, value="↓ Marks per student")
                c.fill      = s["meta_label_fill"]
                c.font      = s["meta_hint_font"]
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border    = s["top_border"]

                current_row += 1

                # ── Student rows ──────────────────────────────────────────────
                for s_idx, (student_id, student_name) in enumerate(students):
                    ws.row_dimensions[current_row].height = 18
                    is_last  = (s_idx == len(students) - 1)
                    row_bdr  = s["bot_border"] if is_last else s["border"]
                    base_fill = s["student_fill"] if s_idx % 2 == 0 else s["alt_student_fill"]

                    # A: test no
                    c = ws.cell(row=current_row, column=1, value=test_no)
                    c.fill = base_fill; c.font = s["id_font"]
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.border = row_bdr

                    # B: admission no
                    c = ws.cell(row=current_row, column=2, value=student_id)
                    c.fill = base_fill; c.font = s["id_font"]
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.border = row_bdr

                    # C: student name
                    c = ws.cell(row=current_row, column=3, value=student_name)
                    c.fill = base_fill; c.font = s["student_font"]
                    c.alignment = Alignment(horizontal='left', vertical='center')
                    c.border = row_bdr

                    # D, E, F: locked (parser reads from META row)
                    for col in (4, 5, 6):
                        c = ws.cell(row=current_row, column=col, value="")
                        c.fill = base_fill; c.border = row_bdr

                    # G: Marks — green, editable
                    c = ws.cell(row=current_row, column=7, value="")
                    c.fill      = s["marks_fill"]
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.border    = row_bdr

                    current_row += 1

                # Blank gap row between tests
                current_row += 1

        # ── Instructions sheet ───────────────────────────────────────────────
        iws = wb.create_sheet("Instructions")
        iws.column_dimensions['A'].width = 72

        if multi_template:
            lines = [
                ("Unit Test Bulk Template — How to Fill", True, 14),
                ("", False, 11),
                ("COLOUR GUIDE", True, 11),
                ("  • Dark blue row (TEST N)  →  One per test. Fill Date, Subject, Unit Name here ONLY.", False, 10),
                ("  • Yellow cells (D, E, F)  →  Type date (YYYY-MM-DD), subject and topic for that test.", False, 10),
                ("  • Green cells (G / Marks) →  Type each student's marks. Leave blank if absent.", False, 10),
                ("  • Grey/white rows         →  Student rows — do NOT edit columns A, B, C, D, E, F.", False, 10),
                ("", False, 11),
                ("STEP-BY-STEP", True, 11),
                ("  1. For each TEST N block, click the yellow Date cell → type the exam date.", False, 10),
                ("  2. Click yellow Subject cell → type subject (Maths / Physics / Chemistry).", False, 10),
                ("  3. Click yellow Unit Name cell → type the topic tested.", False, 10),
                ("  4. Fill the green Marks cell for every student in that block.", False, 10),
                ("     Leave blank for absent students — they will be skipped.", False, 10),
                ("  5. Repeat for all TEST blocks.", False, 10),
                ("  6. Save the file and upload it.", False, 10),
                ("", False, 11),
                ("ACCEPTED MARK VALUES", True, 11),
                ("  Numbers (including negatives for negative marking), 'a' or 'A' for absent.", False, 10),
                ("", False, 11),
                ("DO NOT", True, 11),
                ("  ✗  Edit or delete the TEST N rows.", False, 10),
                ("  ✗  Edit columns A, B, C in student rows.", False, 10),
                ("  ✗  Change the sheet name.", False, 10),
                ("", False, 11),
                ("This template was generated specifically for your batch.", False, 10),
            ]
        else:
            lines = [
                ("Unit Test Template — Instructions", True, 14),
                ("", False, 11),
                ("  1. Fill only the Marks column (col C).", False, 10),
                ("  2. Leave Marks blank for absent students.", False, 10),
                ("  3. Do not edit Admission Number or Student Name columns.", False, 10),
                ("  4. Save and upload the file.", False, 10),
            ]

        for r, (text, bold, size) in enumerate(lines, 1):
            cell = iws.cell(row=r, column=1, value=text)
            cell.font = Font(bold=bold, size=size, name="Calibri")
            iws.row_dimensions[r].height = 16

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        filename = (
            f"unit_test_bulk_template_batch_{batch_id}.xlsx"
            if multi_template else
            f"unit_test_template_batch_{batch_id}.xlsx"
        )
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate template: {str(e)}"
        )
    finally:
        if cursor: cursor.close()
        if conn:   conn.close()


# ── Shared style factory ─────────────────────────────────────────────────────

def _mock_styles():
    thin  = Side(style='thin',   color="BFBFBF")
    thick = Side(style='medium', color="1F4E79")
    return {
        "col_hdr_fill":    PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
        "meta_label_fill": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
        "meta_date_fill":  PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "white_fill":      PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        "alt_fill":        PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
        # Per-subject edit fills (META row unit+total cells)
        "meta_edit": {
            "maths":     PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "physics":   PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
            "chemistry": PatternFill(start_color="EAD1DC", end_color="EAD1DC", fill_type="solid"),
            "biology":   PatternFill(start_color="B6D7A8", end_color="B6D7A8", fill_type="solid"),
        },
        # Per-subject marks fills (student rows)
        "marks_fill": {
            "maths":     PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
            "physics":   PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
            "chemistry": PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
            "biology":   PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
        },
        # Per-subject text colours
        "subj_color": {
            "maths":     "375623",
            "physics":   "783F04",
            "chemistry": "4A1942",
            "biology":   "274E13",
        },
        "col_hdr_font":    Font(bold=True, color="FFFFFF", size=10, name="Calibri"),
        "meta_label_font": Font(bold=True, color="1F4E79", size=10, name="Calibri"),
        "meta_hint_font":  Font(color="9E9E9E", size=9,  name="Calibri", italic=True),
        "meta_val_font":   Font(color="7F4F00", size=10, name="Calibri", bold=True),
        "stu_font":        Font(size=10, name="Calibri"),
        "id_font":         Font(size=10, name="Calibri", color="595959"),
        "border":          Border(left=thin, right=thin, top=thin,  bottom=thin),
        "top_border":      Border(left=thin, right=thin, top=thick, bottom=thin),
        "bot_border":      Border(left=thin, right=thin, top=thin,  bottom=thick),
    }


SUBJ_LABELS = {"maths": "Maths", "physics": "Physics", "chemistry": "Chemistry", "biology": "Biology"}
SUBJ_ORDER  = ["maths", "physics", "chemistry", "biology"]
FIXED_COLS_MOCK  = 4  # A=TestNo, B=AdmNo, C=Name, D=Date


def _build_col_layout(active_subjects):
    """Returns (col_headers, col_widths, subject_col_map, total_col) all 1-indexed."""
    headers = ["Test No", "Admission No", "Student Name", "Exam Date\n(YYYY-MM-DD)"]
    widths  = [9, 18, 28, 20]
    subj_col_map = {}   # subj -> {"unit": col, "total": col, "marks": col}
    for i, subj in enumerate(active_subjects):
        label = SUBJ_LABELS[subj]
        base  = FIXED_COLS_MOCK + i * 3 + 1   # 1-indexed
        subj_col_map[subj] = {"unit": base, "total": base + 1, "marks": base + 2}
        headers += [f"{label}\nUnit Names", f"{label}\nTotal Marks", f"{label}\nMarks"]
        widths  += [28, 14, 12]
    total_col = FIXED_COLS_MOCK + len(active_subjects) * 3 + 1
    headers.append("Test\nTotal Marks")
    widths.append(14)
    return headers, widths, subj_col_map, total_col


# ─────────────────────────────────────────────────────────────────────────────
# 1. TEMPLATE GENERATOR
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/exam/template/mock-test/{batch_id}")
async def get_mock_test_template(
    batch_id: int,
    multi_template: bool = Query(False),
    test_count: int = Query(1, ge=1, le=100),
    current_user: dict = Depends(get_current_user)
):
    """
    Generate and download Excel template for mock/monthly test marks entry.

    multi_template=false  →  simple sheet per batch subjects
    multi_template=true   →  clean multi-test sheet with META rows
                              (fill Date/Units/Totals once per test block,
                               then just per-subject Marks per student)
    """
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT student_id, student_name
            FROM student WHERE batch_id = %s
            ORDER BY
                CASE WHEN student_id ~ '^[0-9]+$' THEN 0 ELSE 1 END,
                CASE WHEN student_id ~ '^[0-9]+$' THEN student_id::BIGINT END,
                student_id ASC, student_name ASC
        """, (batch_id,))
        students = cursor.fetchall()

        if not students:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"No students found for batch ID {batch_id}"
            )

        cursor.execute("SELECT subjects FROM batch WHERE batch_id = %s", (batch_id,))
        batch_row = cursor.fetchone()
        batch_subjects   = batch_row[0] if batch_row else None
        active_subjects  = get_batch_mock_subjects(batch_subjects)   # ordered list

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mock Test Marks"

        if not multi_template:
            # ── Simple single-test template (unchanged behaviour) ─────────────
            hfill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            hfont = Font(bold=True, color="FFFFFF", size=12)
            thin  = Side(style='thin')
            bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
            subj_hdrs = {"maths": "Maths Marks", "physics": "Physics Marks",
                         "chemistry": "Chemistry Marks", "biology": "Biology Marks"}
            headers = ['Admission Number', 'Student Name'] + [subj_hdrs[s] for s in active_subjects]
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.fill = hfill; cell.font = hfont
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = bdr
            for row, (sid, sname) in enumerate(students, 2):
                ws.cell(row=row, column=1, value=sid).border = bdr
                ws.cell(row=row, column=2, value=sname).border = bdr
                for idx in range(3, len(headers) + 1):
                    ws.cell(row=row, column=idx, value="").border = bdr
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 35
            for idx in range(3, len(headers) + 1):
                ws.column_dimensions[get_column_letter(idx)].width = 18

        else:
            # ── Clean multi-test template ─────────────────────────────────────
            s = _mock_styles()
            col_headers, col_widths, subj_col_map, total_col = _build_col_layout(active_subjects)

            # Row 1: column headers
            ws.row_dimensions[1].height = 34
            for c, (header, width) in enumerate(zip(col_headers, col_widths), 1):
                cell = ws.cell(row=1, column=c, value=header)
                cell.fill      = s["col_hdr_fill"]
                cell.font      = s["col_hdr_font"]
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border    = s["border"]
                ws.column_dimensions[get_column_letter(c)].width = width

            ws.freeze_panes = "A2"

            current_row = 2
            for test_no in range(1, test_count + 1):

                # ── META row ──────────────────────────────────────────────────
                ws.row_dimensions[current_row].height = 26

                # A: TEST N label
                c = ws.cell(row=current_row, column=1, value=f"TEST {test_no}")
                c.fill = s["meta_label_fill"]; c.font = s["meta_label_font"]
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = s["top_border"]

                # B, C: blank blue
                for col in (2, 3):
                    c = ws.cell(row=current_row, column=col, value="")
                    c.fill = s["meta_label_fill"]; c.border = s["top_border"]

                # D: yellow date
                c = ws.cell(row=current_row, column=4, value="e.g. 2026-03-07")
                c.fill = s["meta_date_fill"]
                c.font = Font(color="BFBFBF", size=9, name="Calibri", italic=True)
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = s["top_border"]

                # Per subject
                for subj in active_subjects:
                    cols = subj_col_map[subj]
                    color    = s["subj_color"][subj]
                    edit_fill= s["meta_edit"][subj]

                    # Unit names (editable)
                    c = ws.cell(row=current_row, column=cols["unit"], value="")
                    c.fill = edit_fill
                    c.font = Font(size=9, name="Calibri", bold=True, color=color)
                    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    c.border = s["top_border"]

                    # Total marks (editable)
                    c = ws.cell(row=current_row, column=cols["total"], value=100)
                    c.fill = edit_fill
                    c.font = Font(size=10, name="Calibri", bold=True, color=color)
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.border = s["top_border"]

                    # Marks hint
                    c = ws.cell(row=current_row, column=cols["marks"], value="↓")
                    c.fill = s["meta_label_fill"]; c.font = s["meta_hint_font"]
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.border = s["top_border"]

                # Test total hint
                c = ws.cell(row=current_row, column=total_col, value="↓")
                c.fill = s["meta_label_fill"]; c.font = s["meta_hint_font"]
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = s["top_border"]

                current_row += 1

                # ── Student rows ──────────────────────────────────────────────
                for s_idx, (student_id, student_name) in enumerate(students):
                    ws.row_dimensions[current_row].height = 18
                    is_last   = (s_idx == len(students) - 1)
                    row_bdr   = s["bot_border"] if is_last else s["border"]
                    base_fill = s["white_fill"] if s_idx % 2 == 0 else s["alt_fill"]

                    for col, val, font, align in [
                        (1, test_no,      s["id_font"],  "center"),
                        (2, student_id,   s["id_font"],  "center"),
                        (3, student_name, s["stu_font"], "left"),
                        (4, "",           s["stu_font"], "center"),   # date locked
                    ]:
                        c = ws.cell(row=current_row, column=col, value=val)
                        c.fill = base_fill; c.font = font
                        c.alignment = Alignment(horizontal=align, vertical='center')
                        c.border = row_bdr

                    for subj in active_subjects:
                        cols  = subj_col_map[subj]
                        color = s["subj_color"][subj]

                        # Unit and total: locked (from META)
                        for col in (cols["unit"], cols["total"]):
                            c = ws.cell(row=current_row, column=col, value="")
                            c.fill = base_fill; c.border = row_bdr

                        # Marks: coloured, editable
                        c = ws.cell(row=current_row, column=cols["marks"], value="")
                        c.fill = s["marks_fill"][subj]
                        c.font = Font(size=10, name="Calibri", bold=True, color=color)
                        c.alignment = Alignment(horizontal='center', vertical='center')
                        c.border = row_bdr

                    # Test total: empty
                    c = ws.cell(row=current_row, column=total_col, value="")
                    c.fill = base_fill; c.border = row_bdr

                    current_row += 1

                current_row += 1  # blank gap

        # ── Instructions sheet ────────────────────────────────────────────────
        iws = wb.create_sheet("Instructions")
        iws.column_dimensions['A'].width = 75
        if multi_template:
            lines = [
                ("Mock Test Bulk Template — How to Fill", True, 14),
                ("", False, 11),
                ("COLOUR GUIDE", True, 11),
                ("  • Blue row (TEST N)           → One per test. Fill metadata here ONLY.", False, 10),
                ("  • Yellow cell (Exam Date)     → Type date once per block: YYYY-MM-DD", False, 10),
                ("  • 🟢 Green cells  (Maths)     → Unit name, total marks, then marks per student.", False, 10),
                ("  • 🟠 Orange cells (Physics)   → Unit name, total marks, then marks per student.", False, 10),
                ("  • 🩷 Pink cells   (Chemistry) → Unit name, total marks, then marks per student.", False, 10),
                ("  • White/grey rows             → Student rows — do NOT edit A, B, C, D.", False, 10),
                ("", False, 11),
                ("STEP-BY-STEP", True, 11),
                ("  1. In each TEST N row, fill the yellow date cell (YYYY-MM-DD).", False, 10),
                ("  2. Fill the coloured Unit Name cell for each subject (once per block).", False, 10),
                ("  3. Fill the coloured Total Marks cell for each subject (e.g. 100).", False, 10),
                ("  4. In student rows, fill the coloured Marks cells for each subject.", False, 10),
                ("     Leave blank for absent students — they are skipped.", False, 10),
                ("  5. Repeat for all TEST blocks, then save and upload.", False, 10),
                ("", False, 11),
                ("ACCEPTED MARK VALUES", True, 11),
                ("  Numbers (including negatives for negative marking), blank = absent.", False, 10),
                ("", False, 11),
                ("DO NOT", True, 11),
                ("  ✗  Edit or delete the TEST N rows.", False, 10),
                ("  ✗  Edit columns A, B, C, D in student rows.", False, 10),
                ("  ✗  Change the sheet name.", False, 10),
            ]
        else:
            lines = [
                ("Mock Test Template — Instructions", True, 14),
                ("", False, 11),
                ("  1. Fill marks in the subject columns for each student.", False, 10),
                ("  2. Leave blank for absent students.", False, 10),
                ("  3. Do not edit Admission Number or Student Name columns.", False, 10),
                ("  4. Save and upload the file.", False, 10),
            ]
        for r, (text, bold, size) in enumerate(lines, 1):
            c = iws.cell(row=r, column=1, value=text)
            c.font = Font(bold=bold, size=size, name="Calibri")
            iws.row_dimensions[r].height = 16

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        filename = (
            f"mock_test_bulk_template_batch_{batch_id}.xlsx"
            if multi_template else
            f"mock_test_template_batch_{batch_id}.xlsx"
        )
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate mock test template: {str(e)}"
        )
    finally:
        if cursor: cursor.close()
        if conn:   conn.close()


@app.get("/api/exam/daily-test/student/{student_no}")
async def get_student_daily_tests(student_no: int, current_user: dict = Depends(get_current_user)):
    """
    Get all unit test marks for a specific student
    """
    conn = None
    cursor = None
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Fetch all unit tests for the student
        cursor.execute("""
            SELECT 
                test_id,
                grade,
                branch,
                test_date,
                subject,
                unit_name,
                total_marks,
                subject_total_marks,
                test_total_marks,
                created_at
            FROM daily_test
            WHERE student_no = %s
            ORDER BY test_date DESC, subject, unit_name
        """, (student_no,))
        
        tests = cursor.fetchall()
        
        # Format results
        daily_tests = []
        for test in tests:
            daily_tests.append({
                "test_id": test[0],
                "grade": test[1],
                "branch": test[2],
                "test_date": test[3].isoformat() if test[3] else None,
                "subject": test[4],
                "unit_name": test[5],
                "total_marks": test[6],
                "subject_total_marks": test[7],
                "test_total_marks": test[8],
                "created_at": test[9].isoformat() if test[9] else None
            })
        
        return {
            "student_no": student_no,
            "daily_tests": daily_tests,
            "total_tests": len(daily_tests)
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch unit tests: {str(e)}"
        )
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@app.get("/api/exam/mock-test/student/{student_no}")
async def get_student_mock_tests(student_no: int, current_user: dict = Depends(get_current_user)):
    """
    Get all monthly test marks for a specific student
    """
    conn = None
    cursor = None
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Fetch all monthly tests for the student
        cursor.execute("""
            SELECT 
                test_id,
                grade,
                branch,
                test_date,
                maths_marks,
                physics_marks,
                chemistry_marks,
                biology_marks,
                maths_unit_names,
                physics_unit_names,
                chemistry_unit_names,
                biology_unit_names,
                total_marks,
                maths_total_marks,
                physics_total_marks,
                chemistry_total_marks,
                biology_total_marks,
                test_total_marks,
                created_at
            FROM mock_test
            WHERE student_no = %s
            ORDER BY test_date DESC
        """, (student_no,))
        
        tests = cursor.fetchall()
        
        # Format results
        mock_tests = []
        for test in tests:
            mock_tests.append({
                "test_id": test[0],
                "grade": test[1],
                "branch": test[2],
                "test_date": test[3].isoformat() if test[3] else None,
                "maths_marks": test[4],
                "physics_marks": test[5],
                "chemistry_marks": test[6],
                "biology_marks": test[7],
                "maths_unit_names": test[8],
                "physics_unit_names": test[9],
                "chemistry_unit_names": test[10],
                "biology_unit_names": test[11],
                "total_marks": test[12],
                "maths_total_marks": test[13],
                "physics_total_marks": test[14],
                "chemistry_total_marks": test[15],
                "biology_total_marks": test[16],
                "test_total_marks": test[17],
                "created_at": test[18].isoformat() if test[18] else None
            })
        
        return {
            "student_no": student_no,
            "mock_tests": mock_tests,
            "total_tests": len(mock_tests)
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch monthly tests: {str(e)}"
        )
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@app.get("/api/exam/batch-report/{batch_id}")
async def get_batch_report(batch_id: int, current_user: dict = Depends(get_current_user)):
    """
    Get batch report data: all students with basic details,
    per-student unit/monthly test counts, and batch-level totals.
    """
    conn = None
    cursor = None

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # 1. Fetch batch info
        cursor.execute("""
            SELECT batch_id, batch_name, start_year, end_year, type
            FROM batch WHERE batch_id = %s
        """, (batch_id,))
        batch_row = cursor.fetchone()
        if not batch_row:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Batch with ID {batch_id} not found"
            )

        batch_info = {
            "batch_id": batch_row[0],
            "batch_name": batch_row[1],
            "start_year": batch_row[2],
            "end_year": batch_row[3],
            "type": batch_row[4],
        }

        # 2. Fetch all students in the batch with basic details
        cursor.execute("""
            SELECT
                s.student_no,
                s.student_id,
                s.student_name,
                s.gender,
                s.dob,
                s.community,
                s.grade,
                s.enrollment_year,
                s.course,
                s.branch,
                s.student_mobile,
                s.email
            FROM student s
            WHERE s.batch_id = %s
            ORDER BY
                CASE WHEN s.student_id ~ '^[0-9]+$' THEN 0 ELSE 1 END,
                CASE WHEN s.student_id ~ '^[0-9]+$' THEN s.student_id::BIGINT END,
                s.student_id ASC,
                s.student_name ASC
        """, (batch_id,))
        student_rows = cursor.fetchall()

        student_nos = [r[0] for r in student_rows]

        # 3. Per-student unit test counts
        daily_counts = {}
        if student_nos:
            cursor.execute(f"""
                SELECT
                    dt.student_no,
                    COUNT(DISTINCT (
                        dt.test_date,
                        {normalized_subject_sql('dt.subject')},
                        COALESCE(NULLIF(TRIM(dt.unit_name), ''), 'Unknown')
                    )) as cnt
                FROM daily_test dt
                WHERE dt.student_no = ANY(%s)
                GROUP BY dt.student_no
            """, (student_nos,))
            for row in cursor.fetchall():
                daily_counts[row[0]] = row[1]

        # 4. Per-student monthly test counts
        mock_counts = {}
        if student_nos:
            cursor.execute("""
                SELECT
                    mt.student_no,
                    COUNT(DISTINCT (
                        mt.test_date,
                        COALESCE(mt.maths_unit_names, ARRAY[]::text[]),
                        COALESCE(mt.physics_unit_names, ARRAY[]::text[]),
                        COALESCE(mt.chemistry_unit_names, ARRAY[]::text[]),
                        COALESCE(mt.biology_unit_names, ARRAY[]::text[]),
                        mt.maths_total_marks,
                        mt.physics_total_marks,
                        mt.chemistry_total_marks,
                        mt.biology_total_marks,
                        mt.test_total_marks
                    )) as cnt
                FROM mock_test mt
                WHERE mt.student_no = ANY(%s)
                GROUP BY mt.student_no
            """, (student_nos,))
            for row in cursor.fetchall():
                mock_counts[row[0]] = row[1]

        # 5. Total distinct unit tests conducted for this batch
        total_daily_tests = 0
        if student_nos:
            cursor.execute(f"""
                SELECT COUNT(DISTINCT (
                    dt.test_date,
                    {normalized_subject_sql('dt.subject')},
                    COALESCE(NULLIF(TRIM(dt.unit_name), ''), 'Unknown')
                ))
                FROM daily_test dt
                WHERE dt.student_no = ANY(%s)
            """, (student_nos,))
            total_daily_tests = cursor.fetchone()[0] or 0

        # 6. Total distinct monthly tests conducted for this batch
        total_mock_tests = 0
        if student_nos:
            cursor.execute("""
                SELECT COUNT(DISTINCT (
                    mt.test_date,
                    COALESCE(mt.maths_unit_names, ARRAY[]::text[]),
                    COALESCE(mt.physics_unit_names, ARRAY[]::text[]),
                    COALESCE(mt.chemistry_unit_names, ARRAY[]::text[]),
                    COALESCE(mt.biology_unit_names, ARRAY[]::text[]),
                    mt.maths_total_marks,
                    mt.physics_total_marks,
                    mt.chemistry_total_marks,
                    mt.biology_total_marks,
                    mt.test_total_marks
                ))
                FROM mock_test mt
                WHERE mt.student_no = ANY(%s)
            """, (student_nos,))
            total_mock_tests = cursor.fetchone()[0] or 0

        # 7. Fetch all unit test records for batch students
        daily_tests = []
        if student_nos:
            cursor.execute("""
                SELECT s.student_id, s.student_name, dt.test_date,
                      dt.subject, dt.unit_name, dt.total_marks, dt.subject_total_marks, dt.test_total_marks
                FROM daily_test dt
                JOIN student s ON s.student_no = dt.student_no
                WHERE dt.student_no = ANY(%s)
                ORDER BY
                    CASE WHEN s.student_id ~ '^[0-9]+$' THEN 0 ELSE 1 END,
                    CASE WHEN s.student_id ~ '^[0-9]+$' THEN s.student_id::BIGINT END,
                    s.student_id ASC,
                    s.student_name ASC,
                    dt.test_date,
                    dt.subject,
                    dt.unit_name
            """, (student_nos,))
            for r in cursor.fetchall():
                daily_tests.append({
                    "student_id": r[0],
                    "student_name": r[1],
                    "test_date": r[2].isoformat() if r[2] else None,
                    "subject": r[3],
                    "unit_name": r[4],
                    "total_marks": r[5],
                    "subject_total_marks": r[6],
                    "test_total_marks": r[7],
                })

        # 8. Fetch all monthly test records for batch students
        mock_tests = []
        if student_nos:
            cursor.execute("""
                SELECT s.student_id, s.student_name, mt.test_date,
                       mt.maths_marks, mt.physics_marks,
                      mt.chemistry_marks, mt.biology_marks, mt.total_marks,
                      mt.maths_total_marks, mt.physics_total_marks,
                      mt.chemistry_total_marks, mt.biology_total_marks, mt.test_total_marks
                FROM mock_test mt
                JOIN student s ON s.student_no = mt.student_no
                WHERE mt.student_no = ANY(%s)
                ORDER BY
                    CASE WHEN s.student_id ~ '^[0-9]+$' THEN 0 ELSE 1 END,
                    CASE WHEN s.student_id ~ '^[0-9]+$' THEN s.student_id::BIGINT END,
                    s.student_id ASC,
                    s.student_name ASC,
                    mt.test_date
            """, (student_nos,))
            for r in cursor.fetchall():
                mock_tests.append({
                    "student_id": r[0],
                    "student_name": r[1],
                    "test_date": r[2].isoformat() if r[2] else None,
                    "maths_marks": r[3],
                    "physics_marks": r[4],
                    "chemistry_marks": r[5],
                    "biology_marks": r[6],
                    "total_marks": r[7],
                    "maths_total_marks": r[8],
                    "physics_total_marks": r[9],
                    "chemistry_total_marks": r[10],
                    "biology_total_marks": r[11],
                    "test_total_marks": r[12],
                })

        # Build student list
        students = []
        for row in student_rows:
            sno = row[0]
            sid = row[1]
            students.append({
                "student_id": sid,
                "student_no": sno,
                "student_name": row[2],
                "gender": row[3],
                "dob": row[4].isoformat() if row[4] else None,
                "community": row[5],
                "grade": row[6],
                "enrollment_year": row[7],
                "course": row[8],
                "branch": row[9],
                "student_mobile": row[10],
                "email": row[11],
                "daily_test_count": daily_counts.get(sno, 0),
                "mock_test_count": mock_counts.get(sno, 0),
            })

        return {
            "batch": batch_info,
            "total_students": len(students),
            "total_daily_tests_conducted": total_daily_tests,
            "total_mock_tests_conducted": total_mock_tests,
            "students": students,
            "daily_tests": daily_tests,
            "mock_tests": mock_tests,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate batch report: {str(e)}"
        )
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()



# ─────────────────────────────────────────────────────────────────────────────
# BULK EXCEL UPLOAD — UNIT TESTS
# POST /api/exam/daily-test/batch/{batch_id}/upload-excel
#
# Reads the existing multi_template Excel format:
#   Col A: Test No  | Col B: Admission Number | Col C: Student Name
#   Col D: Exam Date (YYYY-MM-DD) | Col E: Subject | Col F: Topic/Unit Name
#   Col G: Marks | Col H: Subject Total Marks | Col I: Test Total Marks
#
# Groups rows by (Test No, Date, Subject, Unit Name) and calls the existing
# create_daily_test_bulk() handler — no DB logic duplicated.
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/exam/daily-test/batch/{batch_id}/upload-excel", status_code=status.HTTP_201_CREATED)
async def upload_daily_test_excel(
    batch_id: int,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """
    Upload a filled bulk unit-test Excel template and insert all tests at once.

    Reads the clean template format generated by GET /api/exam/template/daily-test/{batch_id}?multi_template=true:
      • META row  (col A = "TEST N"):  Date in D, Subject in E, Unit Name in F
      • Student rows:                  Marks in G  (blank = absent/skip)
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only Excel files (.xlsx, .xls) are accepted."
        )

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not read the Excel file. Make sure it is a valid .xlsx file."
        )

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header row 1

    exam_groups = []        # list of DailyTestBulkItem dicts, in order
    parse_errors = []
    current_meta = None     # holds the current META row's parsed info

    def _safe_int(v):
        try:
            return int(v) if v is not None and str(v).strip() not in ("", "-") else None
        except (ValueError, TypeError):
            return None

    def _parse_date(raw, row_idx):
        if raw is None:
            return None, f"Row {row_idx}: Missing Exam Date in TEST block."
        try:
            if isinstance(raw, (datetime, date)):
                return (raw.date() if isinstance(raw, datetime) else raw), None
            return datetime.strptime(str(raw).strip(), "%Y-%m-%d").date(), None
        except ValueError:
            return None, f"Row {row_idx}: Invalid date '{raw}' — expected YYYY-MM-DD."

    for row_idx, row in enumerate(rows, start=2):
        # Skip completely blank rows
        if not any(cell is not None and str(cell).strip() not in ("", "↓ Marks per student") for cell in row):
            continue

        col_a = str(row[0]).strip() if row[0] is not None else ""

        # ── META row detection ────────────────────────────────────────────────
        if col_a.upper().startswith("TEST"):
            date_raw    = row[3]  # col D
            subject_raw = row[4]  # col E
            unit_raw    = row[5]  # col F

            exam_date, date_err = _parse_date(date_raw, row_idx)
            if date_err:
                parse_errors.append(date_err)
                current_meta = None
                continue

            subject = str(subject_raw).strip() if subject_raw else ""
            unit    = str(unit_raw).strip()    if unit_raw    else ""

            if not subject:
                parse_errors.append(f"Row {row_idx} ({col_a}): Missing Subject — block skipped.")
                current_meta = None
                continue

            current_meta = {
                "examName":         col_a,
                "examDate":         exam_date,
                "subject":          subject,
                "unitName":         unit,
                "totalMarks":       100,      # default; override via Subject Total Marks if you add col
                "subjectTotalMarks": None,
                "testTotalMarks":   None,
                "examType":         "daily test",
                "studentMarks":     [],
            }
            exam_groups.append(current_meta)
            continue

        # ── Student row ───────────────────────────────────────────────────────
        if current_meta is None:
            # Row before any META row — skip
            continue

        student_id = str(row[1]).strip() if row[1] is not None else ""
        if not student_id:
            parse_errors.append(f"Row {row_idx}: Missing Admission Number — skipped.")
            continue

        marks_raw = row[6]  # col G
        marks_str = str(marks_raw).strip() if marks_raw is not None else ""
        # Treat placeholder hint text as empty
        if marks_str.lower() in ("", "↓ marks per student"):
            marks_str = ""

        current_meta["studentMarks"].append(
            DailyTestStudentMark(id=student_id, marks=marks_str if marks_str else None)
        )

    # Remove blocks with no student marks at all
    valid_exams = [g for g in exam_groups if g["studentMarks"]]
    skipped_empty = len(exam_groups) - len(valid_exams)
    if skipped_empty:
        parse_errors.append(f"{skipped_empty} test block(s) had no student marks and were skipped.")

    if not valid_exams:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "message": "No valid unit test data found in the Excel file.",
                "parse_errors": parse_errors,
            }
        )

    bulk_payload = DailyTestBulkCreate(
        batch_id=batch_id,
        exams=[DailyTestBulkItem(**g) for g in valid_exams]
    )
    result = await create_daily_test_bulk(bulk_payload, current_user)
    result["total_tests_parsed"] = len(valid_exams)
    result["parse_errors"]       = parse_errors
    return result


# ─────────────────────────────────────────────────────────────────────────────
# BULK EXCEL UPLOAD — MOCK TESTS
# POST /api/exam/mock-test/batch/{batch_id}/upload-excel
#
# Reads the existing multi_template Excel format:
#   Col A: Test No | Col B: Admission Number | Col C: Student Name
#   Col D: Exam Date (YYYY-MM-DD)
#   Then for each active subject (Maths, Physics, Chemistry, Biology):
#       <Subject> Unit Names | <Subject> Total Marks | <Subject> Marks
#   Last col: Test Total Marks
#
# Groups rows by (Test No, Date) and calls create_mock_test_bulk().
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/exam/mock-test/batch/{batch_id}/upload-excel", status_code=status.HTTP_201_CREATED)
async def upload_mock_test_excel(
    batch_id: int,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """
    Upload a filled bulk mock-test Excel template and insert all tests at once.

    Reads the clean template format:
      • META row  (col A = "TEST N"):  Date in D, per-subject Unit+Total in their columns
      • Student rows:                  per-subject Marks in their columns
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only Excel files (.xlsx, .xls) are accepted."
        )

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not read the Excel file. Make sure it is a valid .xlsx file."
        )

    # Determine active subjects from batch
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT subjects FROM batch WHERE batch_id = %s", (batch_id,))
        batch_row = cursor.fetchone()
        if not batch_row:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Batch {batch_id} not found")
        active_subjects = get_batch_mock_subjects(batch_row[0])
    finally:
        if cursor: cursor.close()
        if conn:   conn.close()

    _, _, subj_col_map, total_col = _build_col_layout(active_subjects)
    # Convert to 0-indexed for row tuple access
    subj_col_map_0 = {
        subj: {k: v - 1 for k, v in cols.items()}
        for subj, cols in subj_col_map.items()
    }
    total_col_0 = total_col - 1

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header

    exam_groups  = []
    parse_errors = []
    current_meta = None

    SKIP_VALS = {"", "↓", "↓ marks per student", "none"}

    def _str(v):
        return str(v).strip() if v is not None else ""

    def _safe_int(v):
        try:
            return int(v) if v is not None and _str(v) not in ("", "-") else None
        except (ValueError, TypeError):
            return None

    def _parse_date(raw, row_idx):
        if raw is None:
            return None, f"Row {row_idx}: Missing Exam Date."
        try:
            if isinstance(raw, (datetime, date)):
                return (raw.date() if isinstance(raw, datetime) else raw), None
            return datetime.strptime(_str(raw), "%Y-%m-%d").date(), None
        except ValueError:
            return None, f"Row {row_idx}: Invalid date '{raw}' — expected YYYY-MM-DD."

    for row_idx, row in enumerate(rows, start=2):
        # Skip blank separator rows
        if not any(
            v is not None and _str(v).lower() not in SKIP_VALS
            for v in row
        ):
            continue

        col_a = _str(row[0]).upper()

        # ── META row ──────────────────────────────────────────────────────────
        if col_a.startswith("TEST"):
            exam_date, date_err = _parse_date(row[3], row_idx)   # col D (0-indexed = 3)
            if date_err:
                parse_errors.append(date_err)
                current_meta = None
                continue

            meta = {
                "examName":          col_a.title(),   # "Test 1" etc.
                "examDate":          exam_date,
                "examType":          "mock test",
                "mathsUnitNames":    "",
                "physicsUnitNames":  "",
                "chemistryUnitNames":"",
                "biologyUnitNames":  "",
                "mathsTotalMarks":   None,
                "physicsTotalMarks": None,
                "chemistryTotalMarks": None,
                "biologyTotalMarks": None,
                "testTotalMarks":    None,
                "studentMarks":      [],
            }

            # Read unit names and totals from META row
            for subj in active_subjects:
                cols    = subj_col_map_0[subj]
                unit_v  = _str(row[cols["unit"]])  if cols["unit"]  < len(row) else ""
                total_v = _safe_int(row[cols["total"]]) if cols["total"] < len(row) else None

                # Ignore hint/placeholder values
                if unit_v.lower() in SKIP_VALS or unit_v.lower().startswith("e.g"):
                    unit_v = ""

                meta[f"{subj}UnitNames"]   = unit_v
                meta[f"{subj}TotalMarks"]  = total_v

            # Test total
            if total_col_0 < len(row):
                meta["testTotalMarks"] = _safe_int(row[total_col_0])

            current_meta = meta
            exam_groups.append(meta)
            continue

        # ── Student row ───────────────────────────────────────────────────────
        if current_meta is None:
            continue

        student_id = _str(row[1])   # col B (0-indexed = 1)
        if not student_id:
            parse_errors.append(f"Row {row_idx}: Missing Admission Number — skipped.")
            continue

        mark_kwargs = {"id": student_id}
        for subj in active_subjects:
            cols     = subj_col_map_0[subj]
            marks_v  = _str(row[cols["marks"]]) if cols["marks"] < len(row) else ""
            if marks_v.lower() in SKIP_VALS:
                marks_v = ""
            # Map to Pydantic field names: mathsMarks, physicsMarks, etc.
            mark_kwargs[f"{subj}Marks"] = marks_v if marks_v else None

        current_meta["studentMarks"].append(MockTestStudentMark(**mark_kwargs))

    # Drop blocks with no student marks
    valid_exams    = [g for g in exam_groups if g["studentMarks"]]
    skipped_empty  = len(exam_groups) - len(valid_exams)
    if skipped_empty:
        parse_errors.append(f"{skipped_empty} test block(s) had no student marks and were skipped.")

    if not valid_exams:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"message": "No valid mock test data found.", "parse_errors": parse_errors}
        )

    bulk_payload = MockTestBulkCreate(
        batch_id=batch_id,
        exams=[MockTestBulkItem(**g) for g in valid_exams]
    )
    result = await create_mock_test_bulk(bulk_payload, current_user)
    result["total_tests_parsed"] = len(valid_exams)
    result["parse_errors"]       = parse_errors
    return result
